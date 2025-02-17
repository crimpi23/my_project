from datetime import datetime
from decimal import Decimal
import time
import os
import logging
import bcrypt
from functools import wraps
import psycopg2
import psycopg2.extras
from flask import (
    Flask, render_template, request, session,
    redirect, url_for, flash, jsonify,
    send_file, get_flashed_messages, g
)
from flask_babel import Babel, gettext as _
from openpyxl import Workbook
from io import BytesIO
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import random
import string
import pandas as pd
import numpy as np
from datetime import datetime

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger('psycopg2')




# Налаштування логування
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__, template_folder='templates')

babel = Babel(app)

# Налаштування Babel
app.config['BABEL_DEFAULT_LOCALE'] = 'uk'
app.config['BABEL_SUPPORTED_LOCALES'] = ['uk', 'en', 'sk']

# Конфігурація доступних мов
LANGUAGES = {
    'uk': 'Українська',
    'en': 'English',
    'sk': 'Slovak'
}

def get_locale():
    return session.get('language', app.config['BABEL_DEFAULT_LOCALE'])

babel.init_app(app, locale_selector=get_locale)

# генерація коду для відстеження артикулів в документах
def generate_tracking_code():
    # Використовуємо великі літери та цифри
    chars = string.ascii_uppercase + string.digits
    # Генеруємо 8-значний код
    tracking_code = ''.join(random.choices(chars, k=8))
    return tracking_code

# генерація tracking_code в момент створення нового замовлення постачальнику
def create_supplier_order_details(order_id, article, quantity, order_detail_id):  # Додаємо параметр
    conn = get_db_connection()
    cur = conn.cursor()

    tracking_code = generate_tracking_code()

    cur.execute("""
        INSERT INTO supplier_order_details 
        (supplier_order_id, article, quantity, tracking_code, created_at, order_details_id)  # Додаємо поле
        VALUES (%s, %s, %s, %s, NOW(), %s)  # Додаємо значення
        RETURNING id
    """, (order_id, article, quantity, tracking_code, order_detail_id))  # Додаємо параметр

    detail_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return detail_id


# # Кількість товару в кошику користувача
def get_cart_count(user_id):
    """Повертає кількість товарів у кошику користувача"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(SUM(quantity), 0) 
            FROM cart 
            WHERE user_id = %s
        """, (user_id,))
        count = cursor.fetchone()[0]
        return int(count)
    except Exception as e:
        logging.error(f"Error getting cart count: {e}")
        return 0
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

# Кількість товару в кошику користувача
@app.context_processor
def utility_processor():
    """Додає допоміжні функції до контексту шаблону"""
    def get_user_cart_count():
        user_id = session.get('user_id')
        if user_id:
            return get_cart_count(user_id)
        return 0
    return dict(get_user_cart_count=get_user_cart_count)

# Make LANGUAGES available to all templates
@app.context_processor
def inject_languages():
    return dict(LANGUAGES=LANGUAGES)

@app.route('/set_language/<language>')
def set_language(language):
    session['language'] = language
    return redirect(request.referrer or url_for('index'))

@app.before_request
def before_request():
    g.locale = session.get('language', 'uk')

# Додаємо фільтр для кольорів статусу замовлення
@app.template_filter('status_color')
def status_color(status):
    colors = {
        'new': 'primary',           # Нове замовлення
        'in_review': 'info',        # На розгляді менеджера
        'pending': 'warning',       # В очікуванні
        'accepted': 'success',      # Прийнято
        'cancelled': 'secondary',   # Скасовано
        'ordered_supplier': 'info',      # Замовлено у постачальника
        'invoice_received': 'primary',   # Отримано інвойс
        'in_transit': 'warning',         # В дорозі
        'ready_pickup': 'success',       # Готово до відвантаження
        'completed': 'success'           # Повністю виконано
    }
    return colors.get(status.lower(), 'secondary')


# Секретний ключ для сесій
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# Генерує унікальний токен для користувача
def generate_token():
    return os.urandom(16).hex()


# Хешує пароль для збереження у базі даних
def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


# Перевіряє відповідність пароля хешу з бази
def verify_password(password, stored_hash):
    logging.debug("Verifying password")
    return bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8'))

def get_all_price_list_tables():
    """
    Отримує список усіх таблиць із таблиці price_lists.
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT table_name FROM price_lists;")
            tables = [row[0] for row in cursor.fetchall()]
            logging.debug(f"Fetched tables from price_lists: {tables}")
            return tables
    except Exception as e:
        logging.error(f"Error fetching price list tables: {e}", exc_info=True)
        return []


# Функція для підключення до бази даних
def get_db_connection():
    return psycopg2.connect(
        dsn=os.environ.get('DATABASE_URL'),
        sslmode="require",
        cursor_factory=psycopg2.extras.DictCursor
    )

# Запит про токен / Перевірка токена / Декоратор для перевірки токена
def requires_token_and_roles(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            token = request.view_args.get('token')
            if not token:
                token = request.args.get('token')

            if not token or not validate_token(token):
                flash(_("Invalid token."), "error")
                return redirect(url_for('index'))

            session_role = session.get('role')
            if not session_role:
                flash(_("Invalid token or role."), "error")
                return redirect(url_for('index'))

            role_name = session_role
            if allowed_roles and role_name not in allowed_roles:
                flash(_("Invalid token or role."), "error")
                return redirect(url_for('index'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator


#  отримання даних з selection_buffer 
def get_selection_buffer(user_id):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
            cursor.execute("""
                SELECT article, price, table_name, quantity
                FROM selection_buffer
                WHERE user_id = %s
                ORDER BY article, price;
            """, (user_id,))
            return cursor.fetchall()

# функція націнки для користувачів
def calculate_price(base_price, markup_percentage):
    # Розрахунок кінцевої ціни з урахуванням націнки
    markup_multiplier = Decimal(1) + (Decimal(markup_percentage) / Decimal(100))
    return round(base_price * markup_multiplier, 2)



def get_markup_by_role(role_name):
    """
    Отримання націнки за роллю.
    """
    # Приклад: націнки для ролей
    role_markup_mapping = {
        "admin_user": 0,   # без націнки
        "manager_user": 35,
        "user_29": 29,
        "user_25": 25,
    }
    return role_markup_mapping.get(role_name, 35)  # стандартна націнка 35%


# Функція для перевірки токена
def validate_token(token):
    logging.debug(f"Validating token: {token}")
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT u.id AS user_id, r.name AS role
            FROM tokens t
            JOIN users u ON t.user_id = u.id
            JOIN user_roles ur ON ur.user_id = u.id
            JOIN roles r ON ur.role_id = r.id
            WHERE t.token = %s
        """, (token,))
        result = cursor.fetchone()

        if result:
            session['user_id'] = result['user_id']
            session['role'] = result['role']  # зберігаємо тільки назву ролі
            logging.debug(f"User ID: {result['user_id']}, Role: {result['role']}")

        conn.close()
        return result if result else None
    except Exception as e:
        logging.error(f"Error validating token: {e}")
        return None


def send_email(to_email, subject, ordered_items, missing_articles):
    """
    Відправляє повідомлення на вказану електронну адресу з деталями замовлення.
    """
    try:
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        sender_email = os.getenv("SMTP_EMAIL")  # Отримання адреси з environment variables
        sender_password = os.getenv("SMTP_PASSWORD")  # Отримання пароля з environment variables

        if not sender_email or not sender_password:
            raise ValueError("SMTP credentials are not set in environment variables.")

        # Формування тексту повідомлення
        message_body = _("Thank you for your order!\n\nYour order details:\n")

        if ordered_items:
            message_body += _("Ordered items:\n")
            for item in ordered_items:
                message_body += _(
                    "- Article: {article}, "
                    "Price: {price:.2f}, "
                    "Quantity: {quantity}, "
                    "Comment: {comment}\n"
                ).format(
                    article=item['article'],
                    price=item['price'],
                    quantity=item['quantity'],
                    comment=item['comment'] or _('No comment')
                )

        if missing_articles:
            message_body += _("\nMissing articles:\n")
            for article in missing_articles:
                message_body += f"- {article}\n"

        # Формування повідомлення
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = to_email
        message["Subject"] = subject
        message.attach(MIMEText(message_body, "plain"))

        # Відправка через SMTP
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)

        logging.info(f"Email sent successfully to {to_email}")
    except Exception as e:
        logging.error(f"Failed to send email to {to_email}: {e}")


# Функція для визначення розділювача
def detect_delimiter(file_content):
    delimiters = [',', ';', '\t', ' ']
    sample_lines = file_content.splitlines()[:5]
    counts = {delimiter: 0 for delimiter in delimiters}

    for line in sample_lines:
        for delimiter in delimiters:
            counts[delimiter] += line.count(delimiter)

    return max(counts, key=counts.get)



# Функція для визначення розділювача у рядку
def detect_delimiter(line):
    delimiters = ['\t', ' ', ';', ',']
    for delimiter in delimiters:
        if delimiter in line:
            return delimiter
    return '\t'


# Функція для розбору рядка з артикулом
def parse_article_line(line, available_tables):
    delimiter = detect_delimiter(line)
    parts = line.strip().split(delimiter)

    # Базові параметри
    article = parts[0].strip().upper()
    quantity = int(parts[1]) if len(parts) > 1 and parts[1].strip().isdigit() else 1

    # Перевірка третього параметра
    if len(parts) > 2:
        if parts[2].strip().lower() in available_tables:
            specified_table = parts[2].strip().lower()
            comment = parts[3].strip() if len(parts) > 3 else None
        else:
            specified_table = None
            comment = parts[2].strip()
    else:
        specified_table = None
        comment = None

    return {
        'article': article,
        'quantity': quantity,
        'specified_table': specified_table,
        'comment': comment
    }

def send_order_confirmation_email(to_email, order_id, total_price, cart_items):
    try:
        smtp_server = current_app.config['SMTP_SERVER']
        smtp_port = current_app.config['SMTP_PORT']
        smtp_username = current_app.config['SMTP_USERNAME']
        smtp_password = current_app.config['SMTP_PASSWORD']

        # Формування листа
        msg = MIMEMultipart()
        msg['From'] = smtp_username
        msg['To'] = to_email
        msg['Subject'] = f"Order Confirmation #{order_id}"

        # Тіло листа
        items_list = "".join([f"- {item['quantity']}x {item['product_id']} at ${item['price']}\n" for item in cart_items])
        body = f"""
        Dear Customer,

        Thank you for your order #{order_id}!

        Order Summary:
        {items_list}
        Total Price: ${total_price}

        Best regards,
        MySite Team
        """
        msg.attach(MIMEText(body, 'plain'))

        # Підключення до SMTP і надсилання
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)

        logging.info(f"Order confirmation email sent to {to_email} for order #{order_id}")
    except Exception as e:
        logging.error(f"Failed to send email to {to_email}: {e}", exc_info=True)


# Головна сторінка за токеном
@app.route('/<token>/')
def token_index(token):
    role = validate_token(token)
    if not role:
        flash(_("Invalid token."), "error")
        return redirect(url_for('index'))  # Якщо токен недійсний, перенаправляємо на головну

    # Якщо токен валідний, зберігаємо роль і токен у сесії
    session['token'] = token
    session['role'] = role
    return render_template('user/index.html', role=role)

# Головна сторінка
@app.route('/')
def index():
    token = session.get('token')
    logging.debug(f"Session data in index: {dict(session)}")  # Логування стану сесії

    if not token:
        return render_template('user/search/simple_search.html')

    role = session.get('role')
    logging.debug(f"Role in index: {role}")  # Логування ролі користувача

    if role == "admin":
        return redirect(url_for('admin_dashboard', token=token))
    elif role == "user":
        return render_template('user/index.html', role=role)
    else:
        flash(_("Invalid token or role."), "error")
        return redirect(url_for('simple_search'))



# Пошук для користувачів без токену
@app.route('/simple_search', methods=['GET', 'POST'])
def simple_search():
    if request.method == 'POST':
        article = request.form.get('article', '').strip()
        if not article:
            flash(_("Please enter an article for search."), "error")
            return redirect(url_for('simple_search'))

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT table_name FROM price_lists")
            tables = [row['table_name'] for row in cursor.fetchall()]

            results = []
            for table in tables:
                query = f"""
                    SELECT article, price, '{table}' AS table_name
                    FROM {table}
                    WHERE article = %s
                """
                cursor.execute(query, (article,))
                results.extend(cursor.fetchall())

            if results:
                logging.debug(f"Search results for article '{article}': {results}")
            else:
                logging.debug(f"No results found for article '{article}'.")

        except Exception as e:
            logging.error(f"Error in simple_search: {e}", exc_info=True)
            flash(_("An error occurred during the search. Please try again later."), "error")
            results = []

        finally:
            if 'conn' in locals() and conn:
                conn.close()

        return render_template('user/search/simple_search_results.html', results=results)

    return render_template('user/search/simple_search.html')


# Доступ до адмін панелі
@app.route('/<token>/admin', methods=['GET', 'POST'])
def admin_panel(token):
    try:
        logging.debug(f"Token received in admin_panel: {token}")

        # Валідація токена
        role_data = validate_token(token)
        logging.debug(f"Role data after validation: {role_data}")

        if not role_data or role_data['role'] != 'admin':
            logging.warning(f"Access denied for token: {token}, Role data: {role_data}")
            flash("Access denied. Admin rights are required.", "error")
            return redirect(url_for('simple_search'))

        # Збереження даних сесії
        session['token'] = token
        session['user_id'] = role_data['user_id']
        session['role'] = role_data['role']
        logging.debug(f"Session after saving role: {dict(session)}")

        # Обробка POST-запиту
        if request.method == 'POST':
            password = request.form.get('password')
            logging.debug(f"Password entered: {'******' if password else 'None'}")

            if not password:
                flash("Password is required.", "error")
                return redirect(url_for('admin_panel', token=token))

            conn = get_db_connection()
            cursor = conn.cursor()

            # Перевірка пароля адміністратора
            cursor.execute("""
                SELECT password_hash 
                FROM users
                WHERE id = %s
            """, (role_data['user_id'],))
            admin_password_hash = cursor.fetchone()
            logging.debug(f"Fetched admin password hash: {admin_password_hash}")

            if not admin_password_hash or not verify_password(password, admin_password_hash[0]):
                logging.warning("Invalid admin password attempt.")
                flash("Invalid password.", "error")
                return redirect(url_for('admin_panel', token=token))

            # Встановлення статусу автентифікації
            session['admin_authenticated'] = True
            session.modified = True
            logging.info(f"Admin authenticated for token: {token}")

            return redirect(url_for('admin_dashboard', token=token))

        # Відображення форми входу
        return render_template('admin/admin_login.html', token=token)

    except Exception as e:
        logging.error(f"Error in admin_panel: {e}", exc_info=True)
        flash("An error occurred while accessing the admin panel.", "error")
        return redirect(url_for('simple_search'))

    finally:
        if 'conn' in locals() and conn:
            conn.close()

# Головна сторінка адмінки
@app.route('/<token>/admin/dashboard')
@requires_token_and_roles('admin')
def admin_dashboard(token):
    try:
        logging.debug(f"Session in admin_dashboard: {dict(session)}")  # Логування стану сесії

        if not session.get('admin_authenticated') or session.get('token') != token:
            logging.warning("Access denied. Admin not authenticated or token mismatch.")
            flash("Access denied. Please log in as an admin.", "error")
            return redirect(url_for('admin_panel', token=token))

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT u.id, u.username, r.name AS role_name
            FROM users u
            JOIN user_roles ur ON u.id = ur.user_id
            JOIN roles r ON ur.role_id = r.id
        """)
        users = cursor.fetchall()
        logging.debug(f"Fetched users and roles for dashboard: {users}")  # Логування даних користувачів

        return render_template('admin/dashboard/admin_dashboard.html', users=users, token=token)
    except Exception as e:
        logging.error(f"Error in admin_dashboard: {e}", exc_info=True)
        flash("An error occurred while loading the dashboard.", "error")
        return redirect(url_for('admin_panel', token=token))
    finally:
        if 'conn' in locals() and conn:
            conn.close()



# Створення користувача в адмін панелі:
@app.route('/<token>/admin/create_user', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def create_user(token):
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        email = request.form.get('email')  # Отримання email
        role_id = request.form.get('role_id')

        logging.debug(f"Data received in create_user: username={username}, email={email}, role_id={role_id}")

        if not username or not password or not email or not role_id:
            flash("All fields are required.", "error")
            return redirect(url_for('create_user', token=token))

        hashed_password = hash_password(password)
        user_token = generate_token()

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Додавання нового користувача
            cursor.execute("""
                INSERT INTO users (username, email, password_hash)
                VALUES (%s, %s, %s)
                RETURNING id
            """, (username, email, hashed_password))
            user_id = cursor.fetchone()['id']

            # Призначення ролі
            cursor.execute("""
                INSERT INTO user_roles (user_id, role_id, assigned_at)
                VALUES (%s, %s, NOW())
            """, (user_id, role_id))

            # Додавання токена
            cursor.execute("""
                INSERT INTO tokens (user_id, token)
                VALUES (%s, %s)
            """, (user_id, user_token))

            conn.commit()
            flash(f"User '{username}' created successfully! Token: {user_token}", "success")
        except Exception as e:
            conn.rollback()
            logging.error(f"Error creating user: {e}", exc_info=True)
            flash("Error creating user. Please check logs.", "error")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

        return redirect(url_for('create_user', token=token))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM roles")
    roles = cursor.fetchall()
    conn.close()

    return render_template('admin/users/create_user.html', roles=roles, token=token)


@app.route('/<token>/search', methods=['GET', 'POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def search_articles(token):
    try:
        user_id = session.get('user_id')
        logging.debug(f"Search initiated by user_id={user_id}")

        articles_input = request.form.get('articles', '').strip()
        logging.debug(f"Raw input: {articles_input}")

        if not articles_input:
            logging.warning("Empty articles input")
            flash(_("Please enter at least one article."), "error")
            return redirect(url_for('index'))

        requested_articles = set()
        articles_data = []
        quantities = {}
        comments = {}

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT pl.table_name, pl.brand_id, b.name as brand_name 
                FROM price_lists pl 
                JOIN brands b ON pl.brand_id = b.id
            """)
            available_tables = cursor.fetchall()
            logging.debug(f"Available tables: {available_tables}")

            for line in articles_input.splitlines():
                delimiter = detect_delimiter(line)
                parts = line.strip().split(delimiter)
                logging.debug(f"Processing line: {line}, delimiter: {delimiter}, parts: {parts}")

                if len(parts) == 0:
                    continue

                article = parts[0].strip().upper()
                requested_articles.add(article)
                quantity = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 1

                if len(parts) > 2:
                    if parts[2].strip().lower() in [table[0] for table in available_tables]:
                        specified_table = parts[2].strip().lower()
                        comment = parts[3].strip() if len(parts) > 3 else None
                    else:
                        specified_table = None
                        comment = parts[2].strip()
                else:
                    specified_table = None
                    comment = None

                quantities[article] = quantity
                comments[article] = comment
                articles_data.append({
                    'article': article,
                    'quantity': quantity,
                    'specified_table': specified_table,
                    'comment': comment
                })

            to_add_to_cart = []
            multiple_prices = {}
            found_articles = set()

            user_markup = Decimal(get_markup_percentage(session['user_id']))
            logging.debug(f"User markup: {user_markup}%")

            for item in articles_data:
                article = item['article']
                specified_table = item['specified_table']

                if specified_table:
                    # Логіка для вказаної таблиці
                    if specified_table not in [table[0] for table in available_tables]:
                        logging.warning(f"Specified table {specified_table} not found in available tables")
                        continue

                    cursor.execute(f"SELECT price, brand_id FROM {specified_table} WHERE article = %s", (article,))
                    result = cursor.fetchone()

                    if result:
                        found_articles.add(article)
                        base_price = Decimal(result['price'])
                        final_price = round(base_price * (1 + user_markup / 100), 2)
                        logging.debug(f"Found price for {article} in {specified_table}: base={base_price}, final={final_price}")
                        to_add_to_cart.append({
                            'article': article,
                            'table': specified_table,
                            'price': base_price,
                            'final_price': final_price,
                            'quantity': item['quantity'],
                            'comment': item['comment'],
                            'brand_id': result['brand_id']
                        })
                else:
                    prices_found = []
                    for table, brand_id, brand_name in available_tables:
                        cursor.execute(f"SELECT price FROM {table} WHERE article = %s", (article,))
                        result = cursor.fetchone()
                        if result:
                            found_articles.add(article)
                            base_price = Decimal(result[0])
                            final_price = round(base_price * (1 + user_markup / 100), 2)
                            logging.debug(f"Found price for {article} in {table}: base={base_price}, final={final_price}")
                            prices_found.append({
                                'table_name': table,
                                'base_price': base_price,
                                'price': final_price,
                                'quantity': item['quantity'],
                                'comment': item['comment'],
                                'brand_id': brand_id,
                                'brand_name': brand_name  # Додаємо назву бренду
                            })
                    if prices_found:
                        multiple_prices[article] = prices_found

            # Визначаємо відсутні артикули
            missing_articles = list(requested_articles - found_articles)
            logging.info(f"Missing articles: {missing_articles}")

            if missing_articles:
                flash(_("Articles not found: {articles}").format(
                    articles=', '.join(missing_articles)), "warning")

            # Додаємо в кошик артикули з однією таблицею
            if to_add_to_cart:
                logging.info(f"Adding {len(to_add_to_cart)} items to cart")
                for item in to_add_to_cart:
                    logging.debug(f"Adding to cart: {item}")
                    cursor.execute("""
                        INSERT INTO cart 
                        (user_id, article, table_name, quantity, base_price, final_price, comment, brand_id, added_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON CONFLICT (user_id, article, table_name) 
                        DO UPDATE SET 
                            quantity = cart.quantity + EXCLUDED.quantity,
                            final_price = EXCLUDED.final_price,
                            comment = EXCLUDED.comment,
                            brand_id = EXCLUDED.brand_id
                    """, (
                        session['user_id'],
                        item['article'],
                        item['table'],
                        item['quantity'],
                        item['price'],
                        item['final_price'],
                        item['comment'],
                        item['brand_id']
                    ))
                conn.commit()
                flash(_("Added to cart: {articles}").format(articles=', '.join(item['article'] for item in to_add_to_cart)), "success")

            # Зберігаємо дані в сесії для артикулів з multiple_prices
            session['grouped_results'] = multiple_prices
            session['missing_articles'] = missing_articles
            logging.debug(f"Saved to session - grouped_results: {len(multiple_prices)} items, missing_articles: {len(missing_articles)} items")

            # Якщо є артикули для вибору таблиць
            if multiple_prices:
                logging.info(f"Rendering search results with {len(multiple_prices)} articles having multiple prices")
                return render_template(
                    'user/search/search_results.html',
                    grouped_results=multiple_prices,
                    missing_articles=missing_articles,
                    quantities=quantities,
                    comments=comments,
                    token=token
                )

            # Якщо всі артикули оброблені - перенаправляємо в кошик
            logging.info("All articles processed, redirecting to cart")
            return redirect(url_for('cart', token=token))

    except Exception as e:
        logging.error(f"Error in search_articles: {e}", exc_info=True)
        flash(_("An error occurred while processing your request."), "error")
        return redirect(url_for('index'))


@app.route('/<token>/search_results', methods=['GET', 'POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def search_results(token):
    user_id = session.get('user_id')
    logging.debug(f"Доступ до результатів пошуку для user_id={user_id}")
    
    if not user_id:
        logging.warning("Спроба доступу без автентифікації")
        flash(_("You need to log in to view search results."), "error")
        return redirect(url_for('index'))

    # Якщо це POST-запит, обробляємо вибір користувача
    if request.method == 'POST':
        logging.debug(f"Обробка POST-запиту для user_id={user_id}")
        logging.debug(f"Дані форми: {request.form}")
        
        selected_prices = {}
        grouped_results = session.get('grouped_results', {})
        
        for key, value in request.form.items():
            if key.startswith('selected_price_'):
                article = key.replace('selected_price_', '')
                table_name, price = value.split(':')
                logging.debug(f"Обробка артикула: {article}, таблиця: {table_name}, ціна: {price}")
                
                # Отримуємо базову ціну з grouped_results
                base_price = None
                for option in grouped_results.get(article, []):
                    if option['table_name'] == table_name:
                        base_price = option['base_price']
                        break
                
                logging.debug(f"Знайдена базова ціна: {base_price} для артикула {article}")
                
                try:
                    decimal_price = Decimal(price)
                    logging.debug(f"Конвертовано ціну в Decimal: {decimal_price}")
                    selected_prices[article] = {
                        'table_name': table_name,
                        'base_price': base_price,
                        'price': decimal_price,
                    }
                except Exception as e:
                    logging.error(f"Помилка конвертації ціни для артикула {article}: {e}")
                    continue

        logging.info(f"Оброблено {len(selected_prices)} обраних позицій")
        logging.debug(f"Обрані ціни: {selected_prices}")

        # Зберегти вибір у `selection_buffer`
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Очистити буфер для поточного користувача
            logging.debug(f"Очищення буфера вибору для user_id={user_id}")
            cursor.execute("DELETE FROM selection_buffer WHERE user_id = %s", (user_id,))

            # Додати нові записи
            for article, data in selected_prices.items():
                logging.debug(f"Додавання в selection_buffer: артикул={article}, "
                            f"таблиця={data['table_name']}, базова_ціна={data['base_price']}, "
                            f"фінальна_ціна={data['price']}")
                cursor.execute("""
                    INSERT INTO selection_buffer 
                    (user_id, article, table_name, base_price, price, quantity, added_at)
                    VALUES (%s, %s, %s, %s, %s, 1, NOW())
                """, (user_id, article, data['table_name'], data['base_price'], data['price']))

            conn.commit()
            logging.info(f"Успішно збережено {len(selected_prices)} позицій в буфер")
            flash(_("Your selection has been saved!"), "success")
        except Exception as e:
            conn.rollback()
            logging.error(f"Помилка оновлення буфера вибору: {str(e)}", exc_info=True)
            flash(_("An error occurred while saving your selection. Please try again."), "error")
        finally:
            cursor.close()
            conn.close()
            logging.debug("З'єднання з базою даних закрито")

        return redirect(url_for('search_results', token=token))

    # Якщо це GET-запит, відображаємо результати
    logging.debug(f"Обробка GET-запиту для user_id={user_id}")
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Отримати список прайс-листів
        cursor.execute("SELECT table_name FROM price_lists")
        price_list_tables = [row[0] for row in cursor.fetchall()]

        # Отримати артикул із сесії
        grouped_results = session.get('grouped_results', {})
        if not grouped_results:
            flash(_("No search results found. Please start a new search."), "info")
            return redirect(url_for('index'))

        # Групувати результати по артикулах
        results = {}
        for article, options in grouped_results.items():
            for option in options:
                results.setdefault(article, []).append(option)

        return render_template(
            'user/search/search_results.html',
            grouped_results=results,
            token=token
        )

    except Exception as e:
        logging.error(f"Помилка отримання результатів пошуку: {str(e)}", exc_info=True)
        flash(_("An error occurred while retrieving search results."), "error")
        return redirect(url_for('index'))
    finally:
        cursor.close()
        conn.close()
 
 
@app.route('/<token>/cart', methods=['GET', 'POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def cart(token):
    """
    Функція для обробки кошика користувача:
    - Відображає товари у кошику.
    - Відображає відсутні артикули.
    - Дозволяє оновлювати кількість або видаляти товари.
    """
    try:
        # Отримуємо ID користувача з сесії
        user_id = session.get('user_id')
        if not user_id:
            flash(_("You need to log in to view your cart."), "error")
            logging.warning("Attempt to access cart without user ID.")
            return redirect(url_for('index'))

        logging.debug(f"User ID: {user_id} accessed the cart.")

        # Підключення до бази даних
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # Якщо це POST-запит, обробляємо дії з кошиком (видалення, оновлення кількості)
        if request.method == 'POST':
            action = request.form.get('action')
            article = request.form.get('article')
            table_name = request.form.get('table_name')
            logging.debug(f"POST action received: {action}, article: {article}, table_name: {table_name}")

            if action == 'remove' and article and table_name:
                # Видалення товару з кошика
                cursor.execute(
                    """
                    DELETE FROM cart
                    WHERE user_id = %s AND article = %s AND table_name = %s
                    """,
                    (user_id, article, table_name)
                )
                conn.commit()
                flash("Item removed from cart.", "success")
                logging.info(f"Article {article} removed from cart for user_id {user_id}.")

            elif action == 'update_quantity' and article and table_name:
                # Оновлення кількості товару
                try:
                    new_quantity = int(request.form.get('quantity', 1))
                    if new_quantity > 0:
                        cursor.execute(
                            """
                            UPDATE cart
                            SET quantity = %s
                            WHERE user_id = %s AND article = %s AND table_name = %s
                            """,
                            (new_quantity, user_id, article, table_name)
                        )
                        conn.commit()
                        flash("Item quantity updated.", "success")
                        logging.info(f"Article {article} quantity updated to {new_quantity} for user_id {user_id}.")
                    else:
                        flash("Quantity must be greater than zero.", "error")
                        logging.warning(f"Invalid quantity {new_quantity} provided for article {article}.")
                except ValueError:
                    flash("Invalid quantity provided.", "error")
                    logging.error(f"Non-integer quantity provided for article {article}.")

        # Отримуємо товари з кошика
        cursor.execute("""
            SELECT 
                c.article, 
                c.table_name,
                COALESCE(c.base_price, 0) AS base_price,
                COALESCE(c.final_price, 0) AS final_price,
                c.quantity,
                ROUND(COALESCE(c.final_price, 0) * quantity, 2) AS total_price,
                c.comment,
                b.name AS brand_name  -- Отримуємо назву бренду
            FROM cart c
            JOIN price_lists pl ON c.table_name = pl.table_name
            JOIN brands b ON pl.brand_id = b.id
            WHERE c.user_id = %s
            ORDER BY c.added_at ASC, c.article ASC, c.table_name ASC
        """, (user_id,))

        cart_items = []
        for row in cursor.fetchall():
            cart_items.append({
                'article': row['article'],
                'table_name': row['table_name'],
                'base_price': float(row['base_price']),
                'final_price': float(row['final_price']),
                'quantity': row['quantity'],
                'total_price': float(row['total_price']),
                'comment': row['comment'],
                'brand_name': row['brand_name']  # Передаємо назву бренду
            })
        logging.debug(f"Cart items fetched: {cart_items}")

        # Отримання відсутніх артикулів із сесії
        missing_articles = session.get('missing_articles', [])
        logging.debug(f"Missing articles from session before processing: {missing_articles}")
        if missing_articles:
            missing_articles = list(set(missing_articles))  # Уникнення повторень
        logging.debug(f"Unique missing articles after processing: {missing_articles}")

        # Підрахунок загальної суми
        total_price = sum(item['total_price'] for item in cart_items)
        logging.debug(f"Total price calculated: {total_price}. Preparing to render cart page.")

        # Логування для рендерингу
        logging.info(f"Rendering cart for user_id={user_id}. Cart items count: {len(cart_items)}, Missing articles count: {len(missing_articles)}")
        
        # Додаємо логування для brand_name та table_name
        for item in cart_items:
            logging.debug(f"Item in cart: article={item['article']}, table_name={item['table_name']}, brand_name={item['brand_name']}")

        return render_template(
            'user/cart/cart.html',
            cart_items=cart_items,
            total_price=total_price,
            missing_articles=missing_articles,
            token=token
        )

    except Exception as e:
        logging.error(f"Error in cart for user_id={user_id}: {str(e)}", exc_info=True)
        flash(_("Could not load your cart. Please try again."), "error")
        return redirect(url_for('index'))

    finally:
        # Закриваємо курсор і підключення
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        logging.debug("Database connection closed.")



#проміжковий єтап перед search, який минує search_results до cart
@app.route('/<token>/submit_selection', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def submit_selection(token):
    try:
        user_id = session.get('user_id')
        if not user_id:
            flash(_("User not authenticated"), "error")
            return redirect(url_for('index', token=token))

        grouped_results = session.get('grouped_results', {})
        logging.debug(f"Form data: {request.form}")
        logging.debug(f"Grouped results from session: {grouped_results}")

        selected_articles = []
        for article in grouped_results:
            if request.form.get(f"include_{article}") == "on":
                selected_value = request.form.get(f"selected_{article}")
                if not selected_value:
                    continue

                # Split the selected value into table_name and price|brand_id
                table_name, price_brand = selected_value.split(':')
                price, brand_id = price_brand.split('|')
                
                # Find article data in grouped_results
                article_data = None
                for option in grouped_results[article]:
                    if option['table_name'] == table_name:
                        article_data = option
                        base_price = Decimal(str(option['base_price']))  # Convert to string first
                        break

                if article_data is None:
                    continue

                quantity = int(request.form.get(f"quantity_{article}", 1))
                comment = request.form.get(f"comment_{article}", "").strip() or None
                if comment == "None":  # Fix for "None" string
                    comment = None

                # Convert price to Decimal after removing any potential whitespace
                final_price = Decimal(str(price).strip())

                selected_articles.append({
                    'article': article,
                    'table_name': table_name,
                    'base_price': base_price,
                    'final_price': final_price,
                    'quantity': quantity,
                    'comment': comment,
                    'brand_id': brand_id
                })
                logging.debug(f"Added to selection: {selected_articles[-1]}")

        if not selected_articles:
            flash(_("No articles selected"), "warning")
            return redirect(url_for('index', token=token))

        # Add to cart
        with get_db_connection() as conn:
            cursor = conn.cursor()
            for item in selected_articles:
                cursor.execute("""
                    INSERT INTO cart 
                    (user_id, article, table_name, quantity, base_price, final_price, comment, brand_id, added_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    ON CONFLICT (user_id, article, table_name) DO UPDATE SET
                    quantity = cart.quantity + EXCLUDED.quantity,
                    base_price = EXCLUDED.base_price,
                    final_price = EXCLUDED.final_price,
                    comment = EXCLUDED.comment,
                    brand_id = EXCLUDED.brand_id
                """, (
                    user_id,
                    item['article'],
                    item['table_name'],
                    item['quantity'],
                    item['base_price'],
                    item['final_price'],
                    item['comment'],
                    item['brand_id']
                ))
            conn.commit()

        flash(_("Selection successfully submitted!"), "success")
        return redirect(url_for('cart', token=token))

    except Exception as e:
        logging.error(f"Error in submit_selection: {e}", exc_info=True)
        flash(_("An error occurred during submission"), "error")
        return redirect(url_for('index', token=token))
        
# очищення результату пошуку
@app.route('/<token>/clear_search', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def clear_search(token):
    """Очищає результати пошуку з сесії"""
    try:
        # Очищаємо дані пошуку з сесії
        session.pop('grouped_results', None)
        session.pop('missing_articles', None)
        logging.debug("Search results cleared from session")
        
        flash(_("Search results cleared successfully"), "success")
        # Перенаправляємо на головну сторінку з токеном
        return redirect(url_for('token_index', token=token))
        
    except Exception as e:
        logging.error(f"Error clearing search results: {e}", exc_info=True)
        flash(_("Error clearing search results"), "error")
        return redirect(url_for('token_index', token=token))






# Додавання в кошик користувачем
@app.route('/<token>/add_to_cart', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def add_to_cart(token):
    """
    Додає товар до кошика користувача.
    """
    conn = None
    cursor = None
    try:
        # Отримання даних з форми
        article = request.form.get('article')
        price = Decimal(request.form.get('price'))  # Ціна на момент додавання
        quantity = int(request.form.get('quantity'))  # Кількість товару
        table_name = request.form.get('table_name')
        comment = request.form.get('comment', None)  # Коментар користувача
        brand_id = request.form.get('brand_id') # Отримуємо brand_id
        user_id = session.get('user_id')  # ID користувача

        if not user_id:
            flash(_("User is not authenticated. Please log in."), "error")
            return redirect(url_for('index'))

        logging.debug(f"Adding to cart: Article={article}, Price={price}, Quantity={quantity}, Table={table_name}, User={user_id}")

        conn = get_db_connection()
        cursor = conn.cursor()

        # Перевірка, чи товар вже є в кошику
        cursor.execute("""
            SELECT id, quantity FROM cart
            WHERE user_id = %s AND article = %s AND table_name = %s
        """, (user_id, article, table_name))
        existing_cart_item = cursor.fetchone()

        if existing_cart_item:
            # Оновлення кількості товару в кошику
            new_quantity = existing_cart_item['quantity'] + quantity
            cursor.execute("""
                UPDATE cart
                SET quantity = %s, brand_id = %s
                WHERE id = %s
            """, (new_quantity, brand_id, existing_cart_item['id']))
            logging.info(f"Cart updated: Article={article}, New Quantity={new_quantity}, User ID={user_id}")
        else:
            # Додавання нового товару в кошик
            cursor.execute("""
                INSERT INTO cart (user_id, article, table_name, price, quantity, comment, brand_id, added_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            """, (user_id, article, table_name, price, quantity, comment, brand_id))
            logging.info(f"Product added to cart: Article={article}, Quantity={quantity}, User ID={user_id}")

        conn.commit()
        flash(_("Product added to cart!"), "success")
    except Exception as e:
        logging.error(f"Error in add_to_cart: {e}", exc_info=True)
        flash(_("Error adding product to cart."), "error")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            logging.debug("Database connection closed after adding to cart.")

    # Перенаправлення на сторінку результатів пошуку
    return render_template(
        'user/search/search_results.html',
        grouped_results=session.get('grouped_results', {}),
        quantities=session.get('quantities', {}),
        missing_articles=session.get('missing_articles', []),
    )


# Видалення товару з кошика
@app.route('/<token>/remove_from_cart', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def remove_from_cart(token):
    try:
        user_id = session.get('user_id')
        article = request.form.get('article')
        table_name = request.form.get('table_name')

        logging.debug(f"Removing from cart: Article={article}, Table={table_name}, User={user_id}")

        if not all([user_id, article, table_name]):
            flash(_("Missing required information for removal"), "error")
            return redirect(url_for('cart', token=token))

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM cart 
                WHERE user_id = %s 
                AND article = %s 
                AND table_name = %s
                RETURNING id
            """, (user_id, article, table_name))

            deleted = cursor.fetchone()
            conn.commit()

            if deleted:
                flash(_("Article {article} removed successfully").format(article=article), "success")
                logging.info(f"Article {article} removed from cart for user_id={user_id}")
            else:
                flash(_("Item not found in cart"), "warning")
                logging.warning(f"Failed to remove article {article} for user_id={user_id}")

    except Exception as e:
        logging.error(f"Error removing from cart: {e}", exc_info=True)
        flash(_("Error removing item from cart"), "error")

    return redirect(url_for('cart', token=token))


# Оновлення товару в кошику
@app.route('/<token>/update_cart', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def update_cart(token):
    """
    Оновлює кількість товарів у кошику.
    """
    try:
        # Отримуємо дані з форми
        article = request.form.get('article')
        quantity = request.form.get('quantity')
        user_id = session['user_id']

        if not article or not quantity:
            flash(_("Invalid input: article or quantity missing."), "error")
            logging.error("Missing article or quantity in update_cart form.")
            return redirect(url_for('cart', token=token))

        quantity = int(quantity)

        if quantity < 1:
            flash(_("Quantity must be at least 1."), "error")
            logging.error(f"Invalid quantity: {quantity}")
            return redirect(url_for('cart', token=token))

        logging.debug(f"Updating cart: Article={article}, Quantity={quantity}, User={user_id}")

        conn = get_db_connection()
        cursor = conn.cursor()

        # Оновлення кількості у кошику
        cursor.execute("""
            UPDATE cart
            SET quantity = %s
            WHERE user_id = %s AND article = %s
        """, (quantity, user_id, article))

        conn.commit()
        flash(_("Cart updated successfully!"), "success")

    except Exception as e:
        logging.error(f"Error updating cart: {e}", exc_info=True)
        flash(_("Error updating cart."), "error")
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
        logging.debug("Database connection closed after updating cart.")

    return redirect(url_for('cart', token=token))


# Очищення кошика користувача
@app.route('/<token>/cart/clear', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def clear_cart(token):
    try:
        user_id = session.get('user_id')
        if not user_id:
            flash("User not authenticated", "error")
            return redirect(url_for('token_index', token=token))

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM cart WHERE user_id = %s", (user_id,))
            conn.commit()
            flash("Cart cleared successfully", "success")

    except Exception as e:
        logging.error(f"Error clearing cart: {e}", exc_info=True)
        flash("Error clearing cart", "error")

    return redirect(url_for('cart', token=token))


# з кошика в замовлення
@app.route('/<token>/place_order', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def place_order(token):
    """
    Функція для обробки замовлення:
    - Створює замовлення в таблиці `orders`.
    - Додає деталі замовлення до `order_details`.
    - Очищає кошик користувача.
    - Відправляє підтвердження електронною поштою.
    """
    try:
        user_id = session.get('user_id')
        # Отримуємо відсутні артикули з сесії
        missing_articles = session.get('missing_articles', [])
        
        logging.debug(f"Placing order for user_id={user_id}")

        if not user_id:
            flash(_("User is not authenticated. Please log in."), "error")
            logging.error("Attempt to place order by unauthenticated user.")
            return redirect(url_for('index'))

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # Отримання товарів з кошика
        logging.debug("Fetching cart items...")
        cursor.execute("""
            SELECT article, table_name, base_price, final_price as price,
                   quantity, (final_price * quantity) as total_price, comment, brand_id
            FROM cart
            WHERE user_id = %s
        """, (user_id,))
        cart_items = cursor.fetchall()

        if not cart_items:
            flash(_("Your cart is empty!"), "error")
            logging.warning(f"Cart is empty for user_id={user_id}.")
            return redirect(request.referrer or url_for('cart'))

        # Підрахунок загальної суми
        total_price = sum(item['price'] * item['quantity'] for item in cart_items)
        logging.debug(f"Calculated total price for order: {total_price}")

        # Створення запису замовлення
        cursor.execute("""
            INSERT INTO orders (user_id, total_price, order_date, status)
            VALUES (%s, %s, NOW(), %s)
            RETURNING id
        """, (user_id, total_price, "pending"))
        order_id = cursor.fetchone()['id']
        logging.info(f"Order created with id={order_id} for user_id={user_id}")

        # Додавання деталей замовлення
        ordered_items = []
        for cart_item in cart_items:
            # Додаємо запис в таблицю order_details
            cursor.execute("""
                INSERT INTO order_details
                (order_id, article, table_name, base_price, price, quantity, total_price, comment, brand_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                order_id,
                cart_item['article'],
                cart_item['table_name'],
                cart_item['base_price'],
                cart_item['price'],
                cart_item['quantity'],
                cart_item['total_price'],
                cart_item['comment'],
                cart_item['brand_id']
            ))

            # Логуємо успішне додавання
            logging.info(f"Inserted order detail: order_id={order_id}, article={cart_item['article']}")
            
            # Зберігаємо інформацію про замовлений товар
            ordered_items.append({
                "article": cart_item['article'],
                "price": cart_item['price'],
                "quantity": cart_item['quantity'],
                "comment": cart_item['comment'] or "No comment",
                "table_name": cart_item['table_name']
            })

        # Очищення кошика
        logging.debug(f"Clearing cart for user_id={user_id}...")
        cursor.execute("DELETE FROM cart WHERE user_id = %s", (user_id,))

        # Підтвердження замовлення
        conn.commit()

        # Надсилання підтвердження на email
        cursor.execute("SELECT email FROM users WHERE id = %s", (user_id,))
        user_email = cursor.fetchone()

        if user_email and 'email' in user_email and user_email['email']:
            try:
                # send_email(
                #     to_email=user_email['email'],
                #     subject=f"Order Confirmation - Order #{order_id}",
                #     ordered_items=ordered_items,
                #     missing_articles=session.get('missing_articles', [])
                # )
                logging.info(f"Email sent successfully to {user_email['email']}")
            except Exception as email_error:
                logging.error(f"Failed to send email: {email_error}")
                flash(_("Order placed, but we couldn't send a confirmation email."), "warning")

        # Очищаємо список відсутніх позицій
        session['missing_articles'] = []
        session.modified = True

        logging.info(f"Cart cleared and order placed successfully for user_id={user_id}")
        flash(_("Order placed successfully!"), "success")
        return redirect(request.referrer or url_for('cart'))

    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Error placing order for user_id={user_id}: {str(e)}", exc_info=True)
        flash(_("Error placing order: {error}").format(error=str(e)), "error")
        return redirect(request.referrer or url_for('cart'))
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        logging.debug("Database connection closed.")




# Замовлення користувача
@app.route('/<token>/orders', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def orders(token):
    user_id = session.get('user_id')
    if not user_id:
        flash(_("User is not authenticated."), "error")
        return redirect(url_for('index'))

    # Отримуємо фільтри з запиту
    article_filter = request.args.get('article', '').strip()
    status_filter = request.args.get('status', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    # Логування отриманих фільтрів
    logging.debug(f"Received filters: Article: {article_filter}, Status: {status_filter}, Start Date: {start_date}, End Date: {end_date}")

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Початковий запит до таблиці orders
        query = """
        SELECT * FROM orders
        WHERE user_id = %s
        """
        params = [user_id]

        # Якщо фільтр по артикулу
        if article_filter:
            query += """
            AND EXISTS (
                SELECT 1 
                FROM order_details 
                WHERE order_id = orders.id 
                AND article LIKE %s
            )
            """
            params.append(f"%{article_filter}%")

        # Якщо фільтр по статусу
        if status_filter:
            query += " AND status = %s"
            params.append(status_filter)

        # Якщо фільтр по даті початку
        if start_date:
            query += " AND order_date >= %s"
            params.append(start_date)

        # Якщо фільтр по даті кінця
        if end_date:
            query += " AND order_date <= %s"
            params.append(end_date)

        # Додаємо сортування
        query += " ORDER BY order_date DESC"

        # Логування сформованого запиту
        logging.debug(f"Executing query: {query} with params: {params}")

        # Виконання запиту
        cursor.execute(query, params)
        orders = cursor.fetchall()

        logging.debug(f"Orders retrieved for user_id={user_id} with filters: {article_filter}, {status_filter}, {start_date}, {end_date}")

        return render_template('user/orders/orders.html', orders=orders)

    except Exception as e:
        logging.error(f"Error fetching orders: {e}", exc_info=True)
        flash(_("Error fetching orders."), "error")
        return redirect(url_for('orders', token=token))

    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        logging.debug("Database connection closed.")



# Окреме замовлення користувача по id
@app.route('/<token>/order_details/<int:order_id>')
@requires_token_and_roles('user', 'user_25', 'user_29')
def order_details(token, order_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                order_id, 
                article,
                original_article, 
                table_name, 
                COALESCE(price, 0) as price,
                quantity, 
                COALESCE(total_price, 0) as total_price, 
                COALESCE(comment, '') as comment,
                status
            FROM order_details 
            WHERE order_id = %s
        """, (order_id,))
        
        details = cursor.fetchall()
        formatted_details = [
            {
                'article': row[1],
                'original_article': row[2],
                'table_name': row[3],
                'price': float(row[4]),
                'quantity': row[5],
                'total_price': float(row[6]),
                'comment': row[7],
                'status': row[8] or 'new'
            }
            for row in details
        ]

        return render_template('user/orders/order_details.html',
                            token=token,
                            order_id=order_id,
                            details=formatted_details,
                            total_price=sum(item['total_price'] for item in formatted_details))

    except Exception as e:
        logging.error(f"Помилка завантаження деталей замовлення для order_id={order_id}: {e}", exc_info=True)
        flash(_("Error loading order details."), "error")
        return redirect(url_for('orders', token=token))


@app.route('/<token>/admin/orders/<int:order_id>/update_status', methods=['POST'])
@requires_token_and_roles('admin')
def update_order_item_status(token, order_id):
    conn = None  # Ініціалізуємо conn тут
    try:
        data = request.get_json()  # Отримуємо JSON з тіла запиту
        if not data:
            logging.warning("No JSON data received.")
            return jsonify({"error": "Invalid JSON data."}), 400

        user_id = session.get('user_id')
        conn = get_db_connection()
        cursor = conn.cursor()

        logging.info(f"Received data for order {order_id}: {data}")

        valid_statuses = {
            'new', 'in_review', 'pending', 'accepted',
            'ordered_supplier', 'invoice_received', 'in_transit',
            'ready_pickup', 'completed', 'cancelled'
        }

        if not data.get('items'):
            logging.warning("No items found in request data.")
            return jsonify({"error": "No items provided."}), 400

        for item in data.get('items', []):
            detail_id = item.get('id')
            new_status = item.get('status')
            comment = item.get('comment', None)

            if not detail_id or not new_status:
                logging.warning(f"Missing required fields for item: {item}")
                continue

            if new_status not in valid_statuses:
                logging.warning(f"Invalid status provided: {new_status}")
                continue

            cursor.execute("SELECT status FROM order_details WHERE id = %s;", (detail_id,))
            current_status = cursor.fetchone()

            if current_status:
                current_status = current_status[0]
                logging.info(f"Updating item {detail_id}: {current_status} -> {new_status}")

                cursor.execute("""
                    UPDATE order_details
                    SET status = %s, comment = %s
                    WHERE id = %s;
                """, (new_status, comment, detail_id))

                cursor.execute("""
                    INSERT INTO order_changes (order_id, order_detail_id, field_changed, old_value, new_value, comment, changed_by)
                    VALUES (%s, %s, 'status', %s, %s, %s, %s);
                """, (order_id, detail_id, current_status, new_status, comment, user_id))

        conn.commit()

        # Оновлення статусу замовлення
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE status = 'new') AS new_count,
                COUNT(*) FILTER (WHERE status = 'in_review') AS in_review_count,
                COUNT(*) FILTER (WHERE status = 'pending') AS pending_count,
                COUNT(*) FILTER (WHERE status = 'accepted') AS accepted_count,
                COUNT(*) FILTER (WHERE status = 'ordered_supplier') AS ordered_count,
                COUNT(*) FILTER (WHERE status = 'invoice_received') AS invoice_count,
                COUNT(*) FILTER (WHERE status = 'in_transit') AS transit_count,
                COUNT(*) FILTER (WHERE status = 'ready_pickup') AS ready_count,
                COUNT(*) FILTER (WHERE status = 'completed') AS completed_count,
                COUNT(*) AS total_count
            FROM order_details
            WHERE order_id = %s;
        """, (order_id,))
        counts = cursor.fetchone()

        if counts:
            total_items = counts[9]  # total_count

            # Визначаємо загальний статус замовлення
            if counts[8] == total_items:  # Всі completed
                new_order_status = 'completed'
            elif counts[7] == total_items:  # Всі ready_pickup
                new_order_status = 'ready_pickup'
            elif counts[6] == total_items:  # Всі in_transit
                new_order_status = 'in_transit'
            elif counts[5] == total_items:  # Всі invoice_received
                new_order_status = 'invoice_received'
            elif counts[4] == total_items:  # Всі ordered_supplier
                new_order_status = 'ordered_supplier'
            elif counts[3] == total_items:  # Всі accepted
                new_order_status = 'accepted'
            elif counts[2] == total_items:  # Всі pending
                new_order_status = 'pending'
            elif counts[1] == total_items:  # Всі in_review
                new_order_status = 'in_review'
            elif counts[0] == total_items:  # Всі new
                new_order_status = 'new'
            else:
                new_order_status = 'in_progress'

            cursor.execute("UPDATE orders SET status = %s WHERE id = %s;", (new_order_status, order_id))
            conn.commit()
            logging.info(f"Order {order_id} status updated to {new_order_status}.")

        flash("Order and item statuses updated successfully.", "success")
        return jsonify({"message": "Statuses updated successfully."})

    except Exception as e:
        logging.error(f"Error in updating order items for order {order_id}: {e}")
        return jsonify({"error": "An error occurred while processing the request."}), 500
    finally:
        if conn:
            conn.close()
            logging.info(f"Database connection closed for order {order_id}.")

@app.route('/<token>/order-changes', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def view_order_changes(token):
    """Показує історію змін артикулів та цін в замовленнях користувача."""
    try:
        user_id = session.get('user_id')
        logging.info(f"Viewing order changes for user_id: {user_id}")

        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Отримуємо тільки зміни артикулів та цін
            cursor.execute("""
                SELECT 
                    oc.order_id,
                    oc.order_detail_id,
                    oc.field_changed,
                    oc.old_value,
                    oc.new_value,
                    oc.change_date,
                    oc.comment,
                    od.article,
                    od.status
                FROM order_changes oc
                JOIN order_details od ON oc.order_detail_id = od.id
                JOIN orders o ON oc.order_id = o.id
                WHERE o.user_id = %s
                AND oc.field_changed IN ('article', 'price')
                ORDER BY oc.change_date DESC
            """, (user_id,))
            
            changes = cursor.fetchall()
            
            # Форматування значень для відображення
            formatted_changes = []
            for change in changes:
                formatted_change = dict(change)
                if change['field_changed'] == 'price':
                    # Форматуємо ціни до 2 знаків після коми
                    formatted_change['old_value'] = f"{float(change['old_value']):.2f}"
                    formatted_change['new_value'] = f"{float(change['new_value']):.2f}"
                formatted_changes.append(formatted_change)
            
            return render_template(
                'user/orders/order_changes.html',
                changes=formatted_changes,
                token=token
            )
            
    except Exception as e:
        logging.error(f"Error viewing order changes: {e}", exc_info=True)
        flash("Error loading order changes", "error")
        return redirect(url_for('orders', token=token))

@app.route('/<token>/admin/orders/<int:order_id>', methods=['GET'])
@requires_token_and_roles('admin')
def admin_order_details(token, order_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо інформацію про замовлення та користувача
            cursor.execute("""
                SELECT o.*, u.username, u.email,
                       EXISTS (
                           SELECT 1 FROM order_details 
                           WHERE order_id = o.id 
                           AND (status IS NULL OR status = 'new')
                       ) as has_unprocessed_items
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.id = %s
            """, (order_id,))
            
            order = cursor.fetchone()
            if not order:
                flash("Order not found", "error")
                return redirect(url_for('admin_orders', token=token))

            # Отримуємо деталі замовлення
            cursor.execute("""
                SELECT id, article, table_name, price, quantity, 
                       total_price, COALESCE(status, 'new') as status, comment
                FROM order_details
                WHERE order_id = %s
                ORDER BY id
            """, (order_id,))
            
            order_items = cursor.fetchall()

            return render_template(
                'admin/orders/admin_order_details.html',
                order=order,
                order_items=order_items,
                token=token
            )

    except Exception as e:
        logging.error(f"Error fetching order details: {e}", exc_info=True)
        flash("Failed to load order details", "error")
        return redirect(url_for('admin_orders', token=token))
    

@app.route('/<token>/admin/orders', methods=['GET'])
@requires_token_and_roles('admin')
def admin_orders(token):
    """
    Відображення всіх замовлень для адміністратора з фільтрацією по статусу.
    """
    logging.info(f"Starting admin_orders route with token: {token}")
    status_filter = request.args.get('status', '')
    logging.debug(f"Status filter: {status_filter}")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    try:
        base_query = """
            SELECT 
                o.id, 
                o.order_date, 
                o.total_price, 
                o.status,
                u.username
            FROM orders o
            JOIN users u ON o.user_id = u.id
        """

        params = []
        if status_filter:
            base_query += " WHERE o.status = %s"
            params.append(status_filter)

        base_query += " ORDER BY o.order_date DESC"

        logging.debug(f"Executing query: {base_query} with params: {params}")
        cursor.execute(base_query, params)
        orders = cursor.fetchall()

        logging.info(f"Successfully fetched {len(orders)} orders")
        logging.debug(f"First order sample: {orders[0] if orders else 'No orders'}")

        return render_template(
            'admin/orders/admin_orders.html',
            orders=orders,
            token=token,
            current_status=status_filter
        )

    except Exception as e:
        logging.error(f"Error in admin_orders: {str(e)}", exc_info=True)
        flash("Failed to load orders.", "error")
        return redirect(url_for('admin_dashboard', token=token))

    finally:
        cursor.close()
        conn.close()
        logging.info("Database connection closed in admin_orders")

# Ролі користувачеві
@app.route('/<token>/admin/assign_roles', methods=['GET', 'POST'])
@requires_token_and_roles('admin')  # Вкажіть 'admin' або іншу роль
def assign_roles(token):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Отримання користувачів і ролей
    cursor.execute("""
        SELECT users.id AS user_id, users.username, roles.id AS role_id, roles.name AS role_name
        FROM user_roles
        JOIN users ON user_roles.user_id = users.id
        JOIN roles ON user_roles.role_id = roles.id
        ORDER BY users.username;
    """)
    user_roles = cursor.fetchall()

    cursor.execute("SELECT id, username FROM users;")
    users = cursor.fetchall()

    cursor.execute("SELECT id, name FROM roles;")
    roles = cursor.fetchall()

    if request.method == 'POST':
        action = request.form['action']
        user_id = request.form['user_id']
        role_id = request.form['role_id']

        try:
            if action == 'assign':
                cursor.execute("""
                    SELECT * FROM user_roles
                    WHERE user_id = %s AND role_id = %s;
                """, (user_id, role_id))
                if cursor.fetchone():
                    flash("Role already assigned.", "warning")
                else:
                    cursor.execute("""
                        INSERT INTO user_roles (user_id, role_id)
                        VALUES (%s, %s);
                    """, (user_id, role_id))
                    conn.commit()
                    flash("Role assigned successfully.", "success")
            elif action == 'remove':
                cursor.execute("""
                    DELETE FROM user_roles
                    WHERE user_id = %s AND role_id = %s;
                """, (user_id, role_id))
                conn.commit()
                flash("Role removed successfully.", "success")
        except Exception as e:
            conn.rollback()
            logging.error(f"Error assigning/removing role: {e}", exc_info=True)
            flash("Error assigning/removing role.", "error")

        # Перенаправляємо на ту ж сторінку
        return redirect(url_for('assign_roles', token=token))

    conn.close()
    return render_template('admin/users/assign_roles.html', user_roles=user_roles, users=users, roles=roles, token=token)







@app.route('/<token>/admin/supplier-mapping', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def supplier_mapping(token):
    if request.method == 'POST':
        price_list_id = request.form.get('price_list_id')
        supplier_id = request.form.get('supplier_id')

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE price_lists 
                SET supplier_id = %s 
                WHERE id = %s
            """, (supplier_id, price_list_id))
            conn.commit()

        flash("Mapping updated successfully", "success")

    return redirect(url_for('manage_price_lists', token=token))

#
@app.route('/<token>/admin/manage-price-lists', methods=['GET'])
@requires_token_and_roles('admin')
def manage_price_lists(token):
    with get_db_connection() as conn:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cursor.execute("""
            SELECT 
                pl.id,
                pl.table_name,
                s.name as supplier_name,
                s.delivery_time,
                pl.created_at,
                s.id as supplier_id
            FROM price_lists pl
            LEFT JOIN suppliers s ON pl.supplier_id = s.id
            ORDER BY pl.created_at DESC
        """)
        price_lists = cursor.fetchall()

        # Отримуємо список всіх постачальників для випадаючого списку
        cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        suppliers = cursor.fetchall()

    return render_template('admin/price_lists/manage_price_lists.html',
                           price_lists=price_lists,
                           suppliers=suppliers,
                           token=token)


@app.route('/<token>/admin/update-price-list-supplier', methods=['POST'])
@requires_token_and_roles('admin')
def update_price_list_supplier(token):
    try:
        price_list_id = request.form.get('price_list_id')
        supplier_id = request.form.get('supplier_id')

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE price_lists
                SET supplier_id = %s
                WHERE id = %s
            """, (supplier_id, price_list_id))
            conn.commit()

        flash("Постачальника успішно оновлено", "success")

    except Exception as e:
        logging.error(f"Error updating supplier: {e}")
        flash("Помилка при оновленні постачальника", "error")

    return redirect(url_for('manage_price_lists', token=token))


# Завантаження прайсу в Адмінці
@app.route('/<token>/admin/upload_price_list', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def upload_price_list(token):
    """
    Обробляє завантаження прайс-листу.
    - GET: Повертає сторінку завантаження.
    - POST: Обробляє завантаження файлу та записує дані в базу, включаючи brand_id.
    """
    if request.method == 'GET':
        try:
            logging.info(f"Accessing upload page for token: {token}")
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT table_name FROM price_lists;")
            price_lists = cursor.fetchall()

            # Отримуємо список брендів для вибору
            cursor.execute("SELECT id, name FROM brands;")
            brands = cursor.fetchall()

            conn.close()
            logging.info(f"Fetched price lists: {price_lists}")
            return render_template('admin/price_lists/upload_price_list.html', price_lists=price_lists,
                                   brands=brands, token=token)
        except Exception as e:
            logging.error(f"Error during GET request: {e}")
            flash("Error loading the upload page.", "error")
            return redirect(url_for('admin_dashboard', token=token))

    if request.method == 'POST':
        try:
            logging.info(f"Starting file upload for token: {token}")
            start_time = time.time()

            # Отримання параметрів форми
            table_name = request.form['table_name']
            new_table_name = request.form.get('new_table_name', '').strip()
            brand_id = request.form.get('brand_id')  # Отримання brand_id
            file = request.files.get('file')

            # Логування вхідних даних
            logging.debug(f"Received table_name: {table_name}, new_table_name: {new_table_name}, brand_id: {brand_id}")

            if not file or file.filename == '':
                flash("No file uploaded or selected.", "error")
                logging.warning("No file uploaded or selected.")
                return redirect(url_for('upload_price_list', token=token))

            # Читання файлу
            file_content = file.read().decode('utf-8', errors='ignore')
            delimiter = detect_delimiter(file_content)
            logging.debug(f"Detected delimiter: {delimiter}")
            reader = csv.reader(io.StringIO(file_content), delimiter=delimiter)

            # Обробка даних з файлу
            data = []
            header_skipped = False
            for row in reader:
                if len(row) < 2:
                    logging.warning(f"Skipping invalid row: {row}")
                    continue
                if not header_skipped and not row[1].replace(',', '').replace('.', '').isdigit():
                    header_skipped = True
                    logging.info(f"Skipped header row: {row}")
                    continue
                try:
                    article = row[0].strip().replace(" ", "").upper()
                    price = float(row[1].replace(",", ".").strip())
                    data.append((article, price))
                except ValueError:
                    logging.warning(f"Skipping row with invalid price: {row}")
                    continue

            logging.info(f"Parsed {len(data)} rows from file. Sample: {data[:5]}")

            # Підключення до бази даних
            conn = get_db_connection()
            cursor = conn.cursor()

            # Логіка для створення нової таблиці
            if table_name == 'new':
                if not new_table_name:
                    flash("New table name is required.", "error")
                    logging.error("New table name was not provided.")
                    return redirect(url_for('upload_price_list', token=token))

                table_name = new_table_name.strip().replace(" ", "_").lower()
                if not re.match(r'^[a-z_][a-z0-9_]*$', table_name):
                    logging.warning(f"Invalid table name: {table_name}")
                    flash("Invalid table name. Only lowercase letters, numbers, and underscores are allowed.", "error")
                    return redirect(url_for('upload_price_list', token=token))

                try:
                    cursor.execute(f"""
                        CREATE TABLE IF NOT EXISTS {table_name} (
                            article TEXT PRIMARY KEY,
                            price NUMERIC,
                            brand_id INTEGER REFERENCES brands(id)
                        );
                    """)
                    cursor.execute("""
                        INSERT INTO price_lists (table_name, brand_id, created_at) VALUES (%s, %s, NOW())
                        ON CONFLICT (table_name) DO UPDATE SET brand_id = %s
                    """, (table_name, brand_id, brand_id))
                    conn.commit()
                    logging.info(f"Created new table: {table_name}")
                except Exception as e:
                    logging.error(f"Error creating new table {table_name}: {e}")
                    flash("Error creating new table. Please check the table name and try again.", "error")
                    return redirect(url_for('upload_price_list', token=token))

            # Перевірка існування таблиці
            try:
                cursor.execute(f"SELECT 1 FROM information_schema.tables WHERE table_name = %s;", (table_name,))
                if cursor.fetchone() is None:
                    flash(f"Table '{table_name}' does not exist. Please try again.", "error")
                    logging.error(f"Table '{table_name}' does not exist.")
                    return redirect(url_for('upload_price_list', token=token))
            except Exception as e:
                logging.error(f"Error checking table existence: {e}")
                flash("An error occurred while verifying the table. Please try again.", "error")
                return redirect(url_for('upload_price_list', token=token))

            # Очищення таблиці
            cursor.execute(f"TRUNCATE TABLE {table_name} RESTART IDENTITY;")
            conn.commit()
            logging.info(f"Truncated table: {table_name}")

            # Завантаження даних у таблицю
            output = io.StringIO()
            for row in data:
                output.write(f"{row[0]},{row[1]},{brand_id}\n")
            output.seek(0)
            cursor.copy_expert(f"COPY {table_name} (article, price, brand_id) FROM STDIN WITH (FORMAT CSV);", output)
            conn.commit()

            cache.delete_memoized(search_articles)
            flash(f"Uploaded {len(data)} rows to table '{table_name}' successfully.", "success")
            logging.info(f"Uploaded {len(data)} rows to table '{table_name}' in {time.time() - start_time:.2f} seconds.")
            return redirect(url_for('upload_price_list', token=token))

        except Exception as e:
            logging.error(f"Error during POST request: {e}")
            flash("An error occurred during upload.", "error")
            return redirect(url_for('upload_price_list', token=token))
        finally:
            if 'conn' in locals() and conn:
                conn.close()
                logging.info("Database connection closed.")

# роут для керування постачальниками
@app.route('/<token>/admin/manage-suppliers', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def manage_suppliers(token):
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            delivery_time = request.form.get('delivery_time')

            logging.info(f"Creating new supplier: {name}, delivery time: {delivery_time} days")

            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO suppliers (name, delivery_time)
                    VALUES (%s, %s)
                    RETURNING id
                """, (name, delivery_time))
                supplier_id = cursor.fetchone()[0]
                conn.commit()

                logging.info(f"Created supplier with ID: {supplier_id}")
                flash(f"Постачальника {name} успішно створено", "success")
                return redirect(url_for('manage_suppliers', token=token))

        except Exception as e:
            logging.error(f"Error creating supplier: {e}")
            flash("Помилка при створенні постачальника", "error")

    with get_db_connection() as conn:
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("""
            SELECT id, name, delivery_time, created_at
            FROM suppliers 
            ORDER BY name
        """)
        suppliers = cursor.fetchall()

    return render_template('admin/suppliers/manage_suppliers.html',
                           suppliers=suppliers,
                           token=token)




# створення замовлення постачальнику
@app.route('/<token>/admin/supplier-orders/create', methods=['POST'])
@requires_token_and_roles('admin')
def create_supplier_order(token):
    try:
        supplier_id = request.form.get('supplier_id')
        order_number = f"SO-{supplier_id}-{datetime.now().strftime('%Y%m%d%H%M')}"
        
        logging.info(f"Creating supplier order. Supplier ID: {supplier_id}, Order number: {order_number}")

        conn = get_db_connection()
        cur = conn.cursor()

        # Створюємо нове замовлення постачальнику
        cur.execute("""
            INSERT INTO supplier_orders (supplier_id, order_number, status)
            VALUES (%s, %s, 'new')
            RETURNING id
        """, (supplier_id, order_number))

        supplier_order_id = cur.fetchone()[0]
        logging.info(f"Created supplier order with ID: {supplier_order_id}")

        # Отримуємо деталі замовлення з order_details
        cur.execute("""
            SELECT id, article, quantity 
            FROM order_details 
            WHERE status = 'accepted' AND table_name IN (
                SELECT table_name 
                FROM price_lists 
                WHERE supplier_id = %s
            )
        """, (supplier_id,))

        details = cur.fetchall()
        logging.info(f"Found {len(details)} order details to process")

        # Додаємо деталі до замовлення постачальнику
        for detail in details:
            detail_id, article, quantity = detail
            tracking_code = generate_tracking_code()
            
            logging.info(f"Processing detail - ID: {detail_id}, Article: {article}, Quantity: {quantity}, Tracking: {tracking_code}")

            try:
                # Додаємо в supplier_order_details з order_details_id
                cur.execute("""
                    INSERT INTO supplier_order_details 
                    (supplier_order_id, article, quantity, tracking_code, order_details_id)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                """, (supplier_order_id, article, quantity, tracking_code, detail_id))
                
                inserted_id = cur.fetchone()[0]
                logging.info(f"Inserted supplier_order_detail with ID: {inserted_id}")

                # Оновлюємо статус в order_details
                cur.execute("""
                    UPDATE order_details 
                    SET status = 'ordered_supplier'
                    WHERE id = %s
                """, (detail_id,))
                
                logging.info(f"Updated order_details status for ID: {detail_id}")

            except Exception as detail_error:
                logging.error(f"Error processing detail {detail_id}: {detail_error}")
                continue

        # Оновлюємо статус замовлення
        cur.execute("""
            UPDATE orders 
            SET status = 'ordered_supplier' 
            WHERE id IN (
                SELECT DISTINCT order_id 
                FROM order_details 
                WHERE status = 'ordered'
            )
        """)

        conn.commit()
        logging.info("Supplier order creation completed successfully")
        flash('Supplier order created successfully', 'success')

    except Exception as e:
        conn.rollback()
        logging.error(f"Error creating supplier order: {str(e)}", exc_info=True)
        flash('Error creating supplier order', 'error')

    finally:
        cur.close()
        conn.close()
        logging.info("Database connection closed")

    return redirect(url_for('list_supplier_orders', token=token))


# Функція прийому інвойсу
@app.route('/<token>/admin/process-invoice', methods=['POST'])
@requires_token_and_roles('admin')
def process_invoice(token):
    """Processes uploaded invoice file and creates invoice records"""
    try:
        if 'file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(url_for('upload_invoice', token=token))
            
        file = request.files['file']
        supplier_id = request.form.get('supplier_id')
        invoice_number = request.form.get('invoice_number')
        
        # Validate required fields
        if not all([file, supplier_id, invoice_number]):
            flash("All fields are required", "error")
            return redirect(url_for('upload_invoice', token=token))

        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate and standardize column names
        required_columns = ['Article', 'Quantity', 'Tracking_Code', 'Price']
        if not all(col in df.columns for col in required_columns):
            # Try to find columns by content if headers are missing
            if len(df.columns) >= 4:
                df.columns = required_columns + [f'Column_{i+4}' for i in range(len(df.columns)-4)]
            else:
                flash("Invalid file format - missing required columns", "error")
                return redirect(url_for('upload_invoice', token=token))

        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Create invoice record
            cursor.execute("""
                INSERT INTO supplier_invoices 
                (invoice_number, supplier_id, status, created_at)
                VALUES (%s, %s, 'new', NOW())
                RETURNING id
            """, (invoice_number, supplier_id))
            
            invoice_id = cursor.fetchone()[0]
            
            # Process each row
            for _, row in df.iterrows():
                try:
                    article = str(row['Article']).strip().upper()
                    quantity = int(row['Quantity'])
                    tracking_code = str(row['Tracking_Code']).strip() if pd.notna(row['Tracking_Code']) else None
                    
                    # Додаємо логування для перевірки ціни
                    price = row['Price']
                    logging.info(f"Article: {article}, Price from Excel: {price}")
                    
                    price = float(price)
                    total_price = price * quantity
                    
                    cursor.execute("""
                        INSERT INTO invoice_details 
                        (invoice_id, article, quantity, tracking_code, price, total_price)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (invoice_id, article, quantity, tracking_code, price, total_price))
                except Exception as row_err:
                    logging.error(f"Error processing row: {row.to_dict()}", exc_info=True)
                    flash(f"Error processing row for article {row['Article']}: {row_err}", "error")
                    conn.rollback()
                    return redirect(url_for('upload_invoice', token=token))
            
            conn.commit()
            flash("Invoice uploaded successfully", "success")
            return redirect(url_for('analyze_invoice', token=token, invoice_id=invoice_id))
            
    except Exception as e:
        logging.error(f"Error processing invoice: {e}", exc_info=True)
        flash("Error processing invoice file", "error")
        return redirect(url_for('upload_invoice', token=token))




# функція перевірки схожості артикулів для приймання інвойсу
def find_similar_articles(article1, article2):
    """Перевіряє схожість артикулів"""
    prefixes = ['K', 'V', 'G', 'J', 'WHT', 'VAG']
    clean_article1 = article1
    clean_article2 = article2
    
    for prefix in prefixes:
        if article1.startswith(prefix):
            clean_article1 = article1[len(prefix):]
        if article2.startswith(prefix):
            clean_article2 = article2[len(prefix):]
    
    return clean_article1 == clean_article2


#  Функція аналізу відповідностей в обробці інвойсу від постачальника
@app.route('/<token>/admin/analyze-invoice/<int:invoice_id>', methods=['GET'])
@requires_token_and_roles('admin')
def analyze_invoice(token, invoice_id):
    logging.info(f"Starting analyze_invoice for invoice_id: {invoice_id}")
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Get invoice details
            logging.info(f"Fetching invoice details for invoice_id: {invoice_id}")
            cursor.execute("""
                SELECT 
                    id.*,
                    si.invoice_number,
                    si.supplier_id,
                    si.status as invoice_status,
                    id.price
                FROM invoice_details id
                JOIN supplier_invoices si ON id.invoice_id = si.id
                WHERE id.invoice_id = %s
                AND (id.status IS NULL OR id.status != 'processed')
            """, (invoice_id,))
            
            invoice_items = cursor.fetchall()
            logging.info(f"Found {len(invoice_items)} unprocessed items in invoice")
            for item in invoice_items:
                logging.debug(f"Invoice item: {dict(item)}")

            full_matches = []
            wrong_articles = []
            excess_quantities = []
            missing_quantities = []
            price_highers = []
            price_lowers = []
            no_tracking_codes = []
            different_names = []
            
            for item in invoice_items:
                logging.info(f"Processing invoice item ID: {item['id']}")
                try:
                    if item['tracking_code']:
                        logging.debug(f"Found tracking code: {item['tracking_code']}")
                        cursor.execute("""
                            SELECT 
                                od.id as order_detail_id,
                                od.article,
                                od.original_article,
                                od.quantity as order_quantity,
                                od.status as order_status,
                                od.order_id,
                                sod.tracking_code,
                                od.base_price,
                                od.price,
                                od.total_price,
                                od.comment,
                                od.table_name
                            FROM supplier_order_details sod
                            JOIN order_details od ON sod.order_details_id = od.id
                            WHERE sod.tracking_code = %s
                            AND od.status = 'ordered_supplier'
                        """, (item['tracking_code'],))
                        
                        order_item = cursor.fetchone()
                        if order_item:
                            logging.debug(f"Found matching order item: {dict(order_item)}")
                            
                            invoice_qty = item['quantity']
                            order_qty = order_item['order_quantity']
                            invoice_price = item['price']
                            order_price = order_item['price']
                            base_price = order_item['base_price']
                            
                            logging.info(f"Comparing quantities - Invoice: {invoice_qty}, Order: {order_qty}")
                            logging.info(f"Comparing prices - Invoice: {invoice_price}, Base: {base_price}")
                            
                            if order_item['article'] == item['article']:
                                if invoice_qty == order_qty and invoice_price == base_price:
                                    logging.info(f"Found full match for article {item['article']}")
                                    full_matches.append({
                                        'invoice_item': item,
                                        'order_item': order_item
                                    })
                                else:
                                    if invoice_qty > order_qty:
                                        logging.info(f"Found excess quantity for article {item['article']}")
                                        excess_quantities.append({
                                            'invoice_item': item,
                                            'order_item': order_item,
                                            'excess_qty': invoice_qty - order_qty
                                        })
                                    elif invoice_qty < order_qty:
                                        logging.info(f"Found missing quantity for article {item['article']}")
                                        missing_quantities.append({
                                            'invoice_item': item,
                                            'order_item': order_item,
                                            'missing_qty': order_qty - invoice_qty
                                        })
                                    
                                    if invoice_price > base_price:
                                        logging.info(f"Found higher price for article {item['article']}")
                                        price_highers.append({
                                            'invoice_item': item,
                                            'order_item': order_item,
                                            'price_diff': invoice_price - base_price
                                        })
                                    elif invoice_price < base_price:
                                        logging.info(f"Found lower price for article {item['article']}")
                                        price_lowers.append({
                                            'invoice_item': item,
                                            'order_item': order_item,
                                            'price_diff': base_price - invoice_price
                                        })
                            else:
                                logging.info(f"Article mismatch - Invoice: {item['article']}, Order: {order_item['article']}")
                                if find_similar_articles(item['article'], order_item['article']):
                                    logging.info(f"Found similar article names")
                                    different_names.append({
                                        'invoice_item': item,
                                        'order_item': order_item
                                    })
                                else:
                                    logging.info(f"Found wrong article")
                                    wrong_articles.append({
                                        'invoice_item': item,
                                        'order_item': order_item
                                    })
                        else:
                            logging.warning(f"No matching order found for tracking code: {item['tracking_code']}")
                    else:
                        logging.warning(f"No tracking code for article: {item['article']}")
                        no_tracking_codes.append({
                            'invoice_item': item
                        })
                except Exception as item_err:
                    logging.error(f"Error processing item {item['id']}: {item_err}", exc_info=True)
                    continue

            # Логування результатів аналізу
            logging.info(f"Analysis results for invoice {invoice_id}:")
            logging.info(f"Full matches: {len(full_matches)}")
            logging.info(f"Wrong articles: {len(wrong_articles)}")
            logging.info(f"Excess quantities: {len(excess_quantities)}")
            logging.info(f"Missing quantities: {len(missing_quantities)}")
            logging.info(f"Price highers: {len(price_highers)}")
            logging.info(f"Price lowers: {len(price_lowers)}")
            logging.info(f"No tracking codes: {len(no_tracking_codes)}")
            logging.info(f"Different names: {len(different_names)}")

            return render_template(
                'admin/invoices/analyze.html',
                token=token,
                invoice_id=invoice_id,
                full_matches=full_matches,
                wrong_articles=wrong_articles,
                excess_quantities=excess_quantities,
                missing_quantities=missing_quantities,
                price_highers=price_highers,
                price_lowers=price_lowers,
                no_tracking_codes=no_tracking_codes,
                different_names=different_names
            )
            
    except Exception as e:
        logging.error(f"Critical error in analyze_invoice: {e}", exc_info=True)
        flash("Error analyzing invoice", "error")
        return redirect(url_for('admin_dashboard', token=token))
    

# Функція обробки невідповідностей
@app.route('/<token>/admin/process-mismatches', methods=['POST'])
@requires_token_and_roles('admin')
def process_mismatches(token):
    """Обробляє невідповідності в інвойсі"""
    logging.info("=== Starting process_mismatches ===")
    try:
        invoice_id = request.form.get('invoice_id')
        logging.info(f"Processing invoice_id: {invoice_id}")

        if not invoice_id:
            logging.error("No invoice_id provided in form data")
            flash("Invoice ID is missing", "error")
            return redirect(url_for('admin_dashboard', token=token))

        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо дані інвойсу
            cursor.execute("""
                SELECT invoice_number, supplier_id, status
                FROM supplier_invoices 
                WHERE id = %s
            """, (invoice_id,))
            invoice_data = cursor.fetchone()
            logging.info(f"Invoice data: {dict(invoice_data) if invoice_data else None}")

            if not invoice_data:
                logging.error(f"Invoice not found for id: {invoice_id}")
                flash("Invoice not found", "error")
                return redirect(url_for('admin_dashboard', token=token))

            # Логуємо всі вхідні дані форми
            form_data = dict(request.form)
            logging.info(f"Received form data: {form_data}")

            # Обробка кожної дії з форми
            for key, value in request.form.items():
                logging.debug(f"Processing form field - Key: {key}, Value: {value}")

                try:
                    if key.startswith('action_'):
                        invoice_detail_id = key.replace('action_', '')
                        action = value
                        logging.info(f"Processing action '{action}' for invoice_detail_id: {invoice_detail_id}")

                        # Логуємо поточний стан деталі інвойсу перед обробкою
                        cursor.execute("""
                            SELECT id.*, sod.order_details_id
                            FROM invoice_details id
                            LEFT JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
                            WHERE id.id = %s
                        """, (invoice_detail_id,))
                        detail_before = cursor.fetchone()
                        logging.info(f"Detail before processing: {dict(detail_before) if detail_before else None}")

                        if action == 'update_article':
                            correct_article = request.form.get('correct_article')
                            update_price = request.form.get('update_price') == 'on'
                            logging.info(f"Update article - New article: {correct_article}, Update price: {update_price}")
                            
                            if not handle_wrong_article(invoice_detail_id, correct_article, update_price, conn, cursor):
                                logging.error(f"Failed to handle wrong article for item {invoice_detail_id}")
                                flash(f"Error handling wrong article for item {invoice_detail_id}", "error")

                        elif action == 'accept_name_change':
                            logging.info(f"Accepting name change for invoice_detail_id: {invoice_detail_id}")
                            if not handle_accept_name_change(invoice_detail_id, conn, cursor):
                                logging.error(f"Failed to handle name change for item {invoice_detail_id}")
                                flash(f"Error handling name change for item {invoice_detail_id}", "error")

                        elif action == 'reduce_order':
                            logging.info(f"Reducing order for invoice_detail_id: {invoice_detail_id}")
                            if not handle_missing_quantity(invoice_detail_id, conn, cursor):
                                logging.error(f"Failed to handle missing quantity for item {invoice_detail_id}")
                                flash(f"Error handling missing quantity for item {invoice_detail_id}", "error")

                        elif action == 'update_price':
                            logging.info(f"Updating price for invoice_detail_id: {invoice_detail_id}")
                            if not handle_price_mismatch(invoice_detail_id, action, conn, cursor):
                                logging.error(f"Failed to handle price mismatch for item {invoice_detail_id}")
                                flash(f"Error handling price mismatch for item {invoice_detail_id}", "error")

                        # Логуємо оновлений стан деталі інвойсу
                        cursor.execute("""
                            SELECT id.*, sod.order_details_id
                            FROM invoice_details id
                            LEFT JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
                            WHERE id.id = %s
                        """, (invoice_detail_id,))
                        detail_after = cursor.fetchone()
                        logging.info(f"Detail after processing: {dict(detail_after) if detail_after else None}")

                except Exception as action_err:
                    logging.error(f"Error processing action for invoice_detail_id {invoice_detail_id}: {action_err}", exc_info=True)
                    flash(f"Error handling action for item {invoice_detail_id}: {str(action_err)}", "error")
                    continue

            conn.commit()
            logging.info("Successfully processed all mismatches")
            flash("Невідповідності успішно оброблено", "success")

    except Exception as e:
        logging.error(f"Critical error in process_mismatches: {e}", exc_info=True)
        flash("Помилка обробки невідповідностей", "error")

    logging.info("=== Finished process_mismatches ===")
    return redirect(url_for('analyze_invoice', token=token, invoice_id=invoice_id))

@app.route('/<token>/admin/confirm-invoice/<int:invoice_id>', methods=['POST'])
@requires_token_and_roles('admin')
def confirm_invoice(token, invoice_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Оновлюємо статус замовлення
            cursor.execute("""
                UPDATE order_details
                SET status = 'invoice_received'
                WHERE id IN (
                    SELECT sod.order_details_id
                    FROM invoice_details id
                    JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
                    WHERE id.invoice_id = %s
                )
                AND status = 'ordered_supplier'
            """, (invoice_id,))
            updated_count = cursor.rowcount
            conn.commit()
            logging.info(f"Updated {updated_count} order details status to 'invoice_received'")

            # Оновлюємо статус деталі інвойсу
            cursor.execute("""
                UPDATE invoice_details 
                SET status = 'processed',
                    processed_at = NOW()
                WHERE invoice_id = %s
            """, (invoice_id,))
            updated_count = cursor.rowcount
            logging.info(f"Updated {updated_count} invoice_detail status for invoice_id: {invoice_id}")

            # Оновлюємо статус інвойсу
            cursor.execute("""
                UPDATE supplier_invoices
                SET status = 'accepted'
                WHERE id = %s
            """, (invoice_id,))
            updated_count = cursor.rowcount
            logging.info(f"Updated {updated_count} supplier_invoices status for id: {invoice_id}")

            conn.commit()
            flash("Інвойс успішно підтверджено", "success")

    except Exception as e:
        logging.error(f"Error confirming invoice: {e}", exc_info=True)
        flash("Помилка підтвердження інвойсу", "error")

    return redirect(url_for('list_invoices', token=token))


def handle_accept_name_change(invoice_detail_id, conn, cursor):
    """
    Обробляє прийняття зміни назви артикула.
    """
    try:
        # Отримуємо інформацію про позицію інвойсу
        cursor.execute("""
            SELECT 
                id.article,
                sod.order_details_id
            FROM invoice_details id
            JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
            WHERE id.id = %s
        """, (invoice_detail_id,))
        
        item = cursor.fetchone()
        
        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        # Оновлюємо артикул в order_details
        cursor.execute("""
            UPDATE order_details 
            SET article = %s
            WHERE id = %s
        """, (item['article'], item['order_details_id']))
        logging.info(f"Updated order_details article for id: {item['order_details_id']} to {item['article']}")

        # Оновлюємо статус деталі інвойсу
        cursor.execute("""
            UPDATE invoice_details 
            SET status = 'processed',
                processed_at = NOW()
            WHERE id = %s
        """, (invoice_detail_id,))
        logging.info(f"Updated invoice_detail status for id: {invoice_detail_id}")

        return True

    except Exception as e:
        logging.error(f"Error handling wrong article: {e}", exc_info=True)
        return False


def handle_excess_quantity(invoice_detail_id, conn, cursor):
    """
    Обробляє ситуацію, коли кількість в інвойсі більша за кількість в замовленні.
    """
    try:
        # Отримуємо інформацію про позицію інвойсу
        cursor.execute("""
            SELECT 
                id.article,
                id.quantity as invoice_quantity,
                sod.order_details_id,
                od.quantity as order_quantity
            FROM invoice_details id
            JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
            JOIN order_details od ON sod.order_details_id = od.id
            WHERE id.id = %s
        """, (invoice_detail_id,))
        item = cursor.fetchone()

        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        # Розраховуємо надлишкову кількість
        excess_quantity = item['invoice_quantity'] - item['order_quantity']

        if excess_quantity <= 0:
            logging.warning(f"No excess quantity for invoice_detail_id: {invoice_detail_id}")
            return False

        # Створюємо новий запис order_details для надлишку
        cursor.execute("""
            INSERT INTO order_details 
            (order_id, article, table_name, price, quantity)
            SELECT order_id, article, table_name, price, %s
            FROM order_details
            WHERE id = %s
        """, (excess_quantity, item['order_details_id']))

        # Оновлюємо статус деталі інвойсу
        cursor.execute("""
            UPDATE invoice_details 
            SET status = 'processed'
            WHERE id = %s
        """, (invoice_detail_id,))

        return True

    except Exception as e:
        logging.error(f"Error handling excess quantity: {e}", exc_info=True)
        return False


def handle_missing_quantity(invoice_detail_id, conn, cursor):
    """
    Обробляє ситуацію, коли кількість в інвойсі менша за кількість в замовленні.
    """
    try:
        # Отримуємо інформацію про позицію інвойсу
        cursor.execute("""
            SELECT 
                id.article,
                id.quantity as invoice_quantity,
                sod.order_details_id,
                od.quantity as order_quantity
            FROM invoice_details id
            JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
            JOIN order_details od ON sod.order_details_id = od.id
            WHERE id.id = %s
        """, (invoice_detail_id,))
        item = cursor.fetchone()

        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        # Розраховуємо залишок кількості
        remaining_quantity = item['order_quantity'] - item['invoice_quantity']

        if remaining_quantity <= 0:
            logging.warning(f"No remaining quantity for invoice_detail_id: {invoice_detail_id}")
            return False

        # Оновлюємо кількість в поточному записі order_details
        cursor.execute("""
            UPDATE order_details 
            SET quantity = %s
            WHERE id = %s
        """, (item['invoice_quantity'], item['order_details_id']))

        # Створюємо новий запис order_details для залишку
        cursor.execute("""
            INSERT INTO order_details 
            (order_id, article, table_name, price, quantity)
            SELECT order_id, article, table_name, price, %s
            FROM order_details
            WHERE id = %s
        """, (remaining_quantity, item['order_details_id']))

        # Оновлюємо статус деталі інвойсу
        cursor.execute("""
            UPDATE invoice_details 
            SET status = 'processed'
            WHERE id = %s
        """, (invoice_detail_id,))

        return True

    except Exception as e:
        logging.error(f"Error handling missing quantity: {e}", exc_info=True)
        return False


def handle_price_mismatch(invoice_detail_id, action, conn, cursor):
    """Обробляє невідповідність ціни в інвойсі з урахуванням націнки клієнта."""
    try:
        logging.info(f"Processing price mismatch for invoice_detail_id: {invoice_detail_id}")

        # Отримуємо інформацію про позицію інвойсу та замовлення
        cursor.execute("""
            SELECT 
                id.price as invoice_price,
                sod.order_details_id,
                od.price as order_price,
                od.base_price as old_base_price,
                o.user_id,
                o.id as order_id
            FROM invoice_details id
            JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
            JOIN order_details od ON sod.order_details_id = od.id
            JOIN orders o ON od.order_id = o.id
            WHERE id.id = %s
        """, (invoice_detail_id,))
        
        item = cursor.fetchone()
        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        # Розрахунок нової ціни з націнкою
        base_price = Decimal(item['invoice_price'])
        old_base_price = Decimal(item['old_base_price'])
        
        # Отримуємо націнку користувача
        cursor.execute("""
            SELECT r.markup_percentage
            FROM user_roles ur
            JOIN roles r ON ur.role_id = r.id
            WHERE ur.user_id = %s
        """, (item['user_id'],))
        
        markup_result = cursor.fetchone()
        if not markup_result:
            logging.error(f"No markup found for user_id: {item['user_id']}")
            return False
        
        markup_percentage = Decimal(markup_result[0])
        final_price = base_price * (1 + markup_percentage / 100)

        if action == 'update_price':
            # Оновлюємо ціни та відмітки про зміни
            cursor.execute("""
                UPDATE order_details 
                SET base_price = %s,
                    price = %s,
                    total_price = %s * quantity,
                    price_changed = true,
                    last_change_type = 'price_updated',
                    change_reason = %s
                WHERE id = %s
            """, (
                base_price, 
                final_price, 
                final_price, 
                f"Price updated from invoice (old: {old_base_price}, new: {base_price})",
                item['order_details_id']
            ))
            
            # Додаємо запис в історію змін
            cursor.execute("""
                INSERT INTO order_changes 
                (order_id, order_detail_id, field_changed, old_value, new_value, changed_by, change_date)
                VALUES (%s, %s, 'price', %s, %s, %s, NOW())
            """, (item['order_id'], item['order_details_id'], item['order_price'], final_price, 'invoice_process')) # Змінено тут

        logging.info(f"Price updated for order_detail_id: {item['order_details_id']}")
        return True

    except Exception as e:
        logging.error(f"Error handling price mismatch: {e}", exc_info=True)
        return False



def handle_no_tracking_code(invoice_detail_id, action, conn, cursor):
    """
    Обробляє ситуацію, коли в інвойсі відсутній код відстеження.
    """
    try:
        # Отримуємо інформацію про позицію інвойсу
        cursor.execute("""
            SELECT 
                id.article,
                id.quantity,
                id.price
            FROM invoice_details id
            WHERE id.id = %s
        """, (invoice_detail_id,))
        item = cursor.fetchone()

        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        if action == 'add_to_warehouse':
            # Додаємо товар на склад
            cursor.execute("""
                INSERT INTO warehouse (article, quantity, base_price, price, table_name, added_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
            """, (item['article'], item['quantity'], item['price'], item['price'], 'unknown'))
            logging.info(f"Added item {item['article']} to warehouse")

        # Оновлюємо статус деталі інвойсу
        cursor.execute("""
            UPDATE invoice_details 
            SET status = 'processed',
                processed_at = NOW()
            WHERE id = %s
        """, (invoice_detail_id,))
        logging.info(f"Updated invoice_detail status for id: {invoice_detail_id}")

        return True

    except Exception as e:
        logging.error(f"Error handling no tracking code: {e}", exc_info=True)
        return False


# Додаткові функції обробки невідповідностей в інвойсі
def handle_wrong_article(invoice_detail_id, correct_article, update_price, conn, cursor):
    """
    Обробляє невідповідність артикулів в інвойсі.
    """
    try:
        logging.info(f"handle_wrong_article called for invoice_detail_id: {invoice_detail_id}, correct_article: {correct_article}, update_price: {update_price}")
        # Отримуємо інформацію про позицію інвойсу
        cursor.execute("""
            SELECT 
                id.article,
                id.quantity as invoice_quantity,
                id.price as invoice_price,
                sod.order_details_id
            FROM invoice_details id
            JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
            WHERE id.id = %s
        """, (invoice_detail_id,))
        
        item = cursor.fetchone()
        
        if not item:
            logging.error(f"Item not found for invoice_detail_id: {invoice_detail_id}")
            return False

        # Оновлюємо артикул в order_details
        cursor.execute("""
            UPDATE order_details 
            SET article = %s
            WHERE id = %s
        """, (correct_article, item['order_details_id']))
        logging.info(f"Updated order_details article for id: {item['order_details_id']} to {correct_article}")

        # Перевіряємо ціну
        invoice_price_per_unit = item['invoice_price'] / item['invoice_quantity']
        price_difference_percentage = abs((invoice_price_per_unit - item['invoice_price']) / item['invoice_price']) * 100

        if update_price and price_difference_percentage <= 1:
            # Оновлюємо ціну в order_details
            cursor.execute("""
                UPDATE order_details 
                SET price = %s,
                    total_price = %s * quantity
                WHERE id = %s
            """, (invoice_price_per_unit, invoice_price_per_unit, item['order_details_id']))
            logging.info(f"Updated order_details price for id: {item['order_details_id']} to {invoice_price_per_unit}")

        # Оновлюємо статус деталі інвойсу
        cursor.execute("""
            UPDATE invoice_details 
            SET status = 'processed',
                processed_at = NOW()
            WHERE id = %s
        """, (invoice_detail_id,))
        logging.info(f"Updated invoice_detail status for id: {invoice_detail_id}")

        return True

    except Exception as e:
        logging.error(f"Error handling wrong article: {e}", exc_info=True)
        return False

# перегляд складу
@app.route('/<token>/admin/warehouse', methods=['GET'])
@requires_token_and_roles('admin')
def warehouse_items(token):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            cursor.execute("""
                SELECT 
                    w.id,
                    w.article,
                    w.quantity,
                    COALESCE(w.base_price, 0) as base_price,
                    COALESCE(w.price, 0) as price,
                    w.table_name,
                    w.added_at,
                    w.invoice_number,
                    si.supplier_id,
                    s.name as supplier_name
                FROM warehouse w
                LEFT JOIN supplier_invoices si ON w.invoice_id = si.id
                LEFT JOIN suppliers s ON si.supplier_id = s.id
                ORDER BY w.added_at DESC
            """)
            
            items = cursor.fetchall()
            
            return render_template(
                'admin/warehouse/list_items.html',
                items=items,
                token=token
            )
            
    except Exception as e:
        logging.error(f"Error fetching warehouse items: {e}")
        flash("Error loading warehouse items", "error")
        return redirect(url_for('admin_dashboard', token=token))


# сторінка для завантаження інвойсу:
@app.route('/<token>/admin/invoice/upload', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def upload_invoice(token):
    if request.method == 'GET':
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Get suppliers list
            cursor.execute("SELECT id, name FROM suppliers")
            suppliers = cursor.fetchall()
            
            return render_template('admin/invoices/upload.html', 
                                 suppliers=suppliers,
                                 token=token)
        except Exception as e:
            logging.error(f"Error in GET: {e}")
            flash("Error loading suppliers", "error")
            return redirect(url_for('admin_dashboard', token=token))
                             
    elif request.method == 'POST':
        try:
            file = request.files.get('invoice_file')
            supplier_id = request.form.get('supplier_id')
            invoice_number = request.form.get('invoice_number')
            
            logging.debug(f"Received file: {file.filename if file else None}")
            logging.debug(f"Supplier ID: {supplier_id}")
            logging.debug(f"Invoice number: {invoice_number}")
            
            if not all([file, supplier_id, invoice_number]):
                flash("All fields are required", "error") 
                return redirect(url_for('upload_invoice', token=token))
                
            # Read Excel file
            df = pd.read_excel(file, dtype={'Price': str})  # Явне вказання типу даних
            logging.debug(f"Excel data shape: {df.shape}")
            logging.debug(f"Columns: {df.columns.tolist()}")
            
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Create invoice record
                cursor.execute("""
                    INSERT INTO supplier_invoices 
                    (invoice_number, supplier_id, status, created_at)
                    VALUES (%s, %s, 'new', NOW())
                    RETURNING id
                """, (invoice_number, supplier_id))
                
                invoice_id = cursor.fetchone()[0]
                
                # Save invoice details
                for _, row in df.iterrows():
                    # Отримуємо значення ціни з файлу Excel
                    logging.info(f"Row data: {row}")  # Логування всього рядка
                    price = row.get('Price')
                    logging.info(f"Article: {row['Article']}, Raw Price from Excel: {price}")

                    # Перевіряємо, чи ціна не є порожньою
                    if price is None:
                        logging.warning(f"Price is empty for Article: {row['Article']}")
                        price_value = 0.0  # або можна вказати інше значення за замовчуванням
                    else:
                        try:
                            price_str = str(price).replace(',', '.')  # Заміна коми на крапку
                            price_value = float(price_str)
                        except Exception as err:
                            logging.error(f"Failed to convert price to float for Article: {row['Article']}, price: {price}", exc_info=True)
                            price_value = 0.0

                    # Логування значення та типу ціни перед виконанням SQL-запиту
                    logging.info(f"Inserting row for Article: {row['Article']} with Price: {price_value} (type: {type(price_value)})")

                    cursor.execute("""
                        INSERT INTO invoice_details 
                        (invoice_id, article, quantity, tracking_code, price)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (invoice_id, row['Article'], 
                          row['Quantity'], row.get('Tracking_Code'), price_value))
                          
                conn.commit()
                
            flash("Invoice uploaded successfully", "success")
            return redirect(url_for('analyze_invoice', 
                                  token=token,
                                  invoice_id=invoice_id))
                                  
        except Exception as e:
            logging.error(f"Error uploading invoice: {e}")
            flash("Error uploading invoice", "error")
            return redirect(url_for('upload_invoice', token=token))


# маршрут для списку інвойсів в адмін-панелі
@app.route('/<token>/admin/invoices', methods=['GET'])
@requires_token_and_roles('admin')
def list_invoices(token):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Отримуємо всі інвойси з інформацією про постачальника
            cursor.execute("""
                SELECT 
                    si.id,
                    si.invoice_number,
                    si.created_at,
                    si.status,
                    s.name as supplier_name,
                    COUNT(id.id) as items_count,
                    COUNT(CASE WHEN id.status = 'processed' THEN 1 END) as processed_count
                FROM supplier_invoices si
                JOIN suppliers s ON si.supplier_id = s.id
                LEFT JOIN invoice_details id ON si.id = id.invoice_id
                GROUP BY si.id, si.invoice_number, si.created_at, si.status, s.name
                ORDER BY si.created_at DESC
            """)
            invoices = cursor.fetchall()
            
            return render_template(
                'admin/invoices/list_invoices.html',
                invoices=invoices,
                token=token
            )
            
    except Exception as e:
        logging.error(f"Error fetching invoices: {e}", exc_info=True)
        flash("Error loading invoices", "error")
        return redirect(url_for('admin_dashboard', token=token))

# функція для перегляду деталей інвойсу
@app.route('/<token>/admin/invoice/<int:invoice_id>/details', methods=['GET'])
@requires_token_and_roles('admin')
def invoice_details(token, invoice_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            
            # Отримуємо інформацію про інвойс
            cursor.execute("""
                SELECT 
                    si.*,
                    s.name as supplier_name
                FROM supplier_invoices si
                JOIN suppliers s ON si.supplier_id = s.id
                WHERE si.id = %s
            """, (invoice_id,))
            
            invoice = cursor.fetchone()
            if not invoice:
                flash("Invoice not found", "error")
                return redirect(url_for('list_invoices', token=token))

            # Отримуємо деталі інвойсу
            cursor.execute("""
                SELECT 
                    id.*,
                    CASE 
                        WHEN sod.order_details_id IS NOT NULL THEN od.article 
                        ELSE NULL 
                    END as original_article,
                    CASE 
                        WHEN sod.order_details_id IS NOT NULL THEN od.quantity 
                        ELSE NULL 
                    END as ordered_quantity,
                    CASE
                        WHEN w.id IS NOT NULL THEN TRUE
                        ELSE FALSE
                    END as in_warehouse,
                    od.base_price,
                    od.price AS customer_price,
                    u.username AS customer_username,
                    u.email AS customer_email
                FROM invoice_details id
                LEFT JOIN supplier_order_details sod ON id.tracking_code = sod.tracking_code
                LEFT JOIN order_details od ON sod.order_details_id = od.id
                LEFT JOIN warehouse w ON id.article = w.article AND id.invoice_id = w.invoice_id
                LEFT JOIN orders o ON od.order_id = o.id
                LEFT JOIN users u ON o.user_id = u.id
                WHERE id.invoice_id = %s
                ORDER BY id.id
            """, (invoice_id,))
            
            invoice_items = cursor.fetchall()

            return render_template(
                'admin/invoices/invoice_details.html',
                invoice=invoice,
                items=invoice_items,
                token=token
            )
            
    except Exception as e:
        logging.error(f"Error fetching invoice details: {e}", exc_info=True)
        flash("Error loading invoice details", "error")
        return redirect(url_for('list_invoices', token=token))

@app.route('/<token>/admin/compare_prices', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def compare_prices(token):
    if request.method == 'GET':
        try:
            # Отримуємо список таблиць із прайсами
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cursor.execute("SELECT table_name FROM price_lists;")
            price_lists = cursor.fetchall()
            conn.close()
            logging.info("Fetched price list tables successfully.")
            return render_template('user/search/compare_prices.html', price_lists=price_lists)
        except Exception as e:
            logging.error(f"Error during GET request: {str(e)}", exc_info=True)
            flash("Failed to load price list tables.", "error")
            return redirect(request.referrer or url_for('admin_panel'))

    if request.method == 'POST':
        try:
            form_data = request.form.to_dict()
            logging.info(f"Form data: {form_data}")

            # Якщо запит на експорт
            if 'export_excel' in request.form:
                logging.info("Export to Excel initiated.")
                data = session.get('comparison_results')
                if not data:
                    logging.error("No data to export!")
                    flash("No data to export!", "error")
                    return redirect(request.referrer or url_for('compare_prices'))
                return export_to_excel(
                    data['better_in_first'],
                    data['better_in_second'],
                    data['same_prices']
                )

            # Обробка даних для порівняння
            articles_input = request.form.get('articles', '').strip()
            selected_prices = request.form.getlist('price_tables')

            if not articles_input or not selected_prices:
                flash("Please enter articles and select price tables.", "error")
                return redirect(request.referrer or url_for('compare_prices'))

            # Розбиваємо артикулі
            articles = [line.strip() for line in articles_input.splitlines() if line.strip()]
            logging.info(f"Articles to compare: {articles}")

            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо ціни з вибраних таблиць
            results = {}
            for table in selected_prices:
                query = f"SELECT article, price FROM {table} WHERE article = ANY(%s);"
                cursor.execute(query, (articles,))
                for row in cursor.fetchall():
                    article = row['article']
                    price = row['price']
                    if article not in results:
                        results[article] = []
                    results[article].append({'price': price, 'table': table})

            conn.close()

            # Порівняння цін
            better_in_first, better_in_second, same_prices = [], [], []
            for article, prices in results.items():
                min_price = min(prices, key=lambda x: x['price'])
                if len([p for p in prices if p['price'] == min_price['price']]) > 1:
                    same_prices.append({
                        'article': article,
                        'price': min_price['price'],
                        'tables': ', '.join(p['table'] for p in prices if p['price'] == min_price['price'])
                    })
                elif min_price['table'] == selected_prices[0]:
                    better_in_first.append({'article': article, **min_price})
                elif min_price['table'] == selected_prices[1]:
                    better_in_second.append({'article': article, **min_price})

            # Зберігаємо результати у сесії
            session['comparison_results'] = {
                'better_in_first': better_in_first,
                'better_in_second': better_in_second,
                'same_prices': same_prices
            }

            logging.info("Comparison completed successfully.")
            return render_template(
                'user/search/compare_prices_results.html',
                better_in_first=better_in_first,
                better_in_second=better_in_second,
                same_prices=same_prices
            )

        except Exception as e:
            logging.error(f"Error during POST request: {str(e)}", exc_info=True)
            flash("An error occurred during comparison.", "error")
            return redirect(request.referrer or url_for('compare_prices'))

# Маршрут для відображення форми (new_supplier_order):
@app.route('/<token>/admin/supplier-orders/new', methods=['GET'])
@requires_token_and_roles('admin')
def new_supplier_order(token):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("SELECT id, name FROM suppliers ORDER BY name")
    suppliers = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('admin/supplier_orders_create.html',
                           token=token,
                           suppliers=suppliers)


@app.route('/<token>/api/price-lists/<supplier_id>')
@requires_token_and_roles('admin')
def get_supplier_price_lists(token, supplier_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("""
        SELECT id, table_name 
        FROM price_lists 
        WHERE supplier_id = %s
    """, (supplier_id,))

    price_lists = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify({'price_lists': price_lists})

# сторінка перегляду замовлень постачальників
@app.route('/<token>/admin/supplier-orders', methods=['GET'])
@requires_token_and_roles('admin')
def list_supplier_orders(token):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Отримуємо всі замовлення з деталями
    cur.execute("""
        SELECT 
            so.id,
            so.order_number,
            so.created_at,
            so.status,
            s.name as supplier_name,
            COUNT(sod.id) as items_count,
            SUM(sod.quantity) as total_quantity
        FROM supplier_orders so
        JOIN suppliers s ON so.supplier_id = s.id
        LEFT JOIN supplier_order_details sod ON so.id = sod.supplier_order_id
        GROUP BY so.id, s.name
        ORDER BY so.created_at DESC
    """)

    orders = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('admin/supplier_orders_list.html',
                           token=token,
                           orders=orders)

# маршрут для перегляду деталей замовлення постачальнику
@app.route('/<token>/admin/supplier-orders/<order_id>', methods=['GET'])
@requires_token_and_roles('admin')
def view_supplier_order(token, order_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Отримуємо основну інформацію про замовлення
    cur.execute("""
        SELECT 
            so.*,
            s.name as supplier_name
        FROM supplier_orders so
        JOIN suppliers s ON so.supplier_id = s.id
        WHERE so.id = %s
    """, (order_id,))

    order = cur.fetchone()

    # Отримуємо деталі замовлення
    cur.execute("""
        SELECT *
        FROM supplier_order_details
        WHERE supplier_order_id = %s
        ORDER BY created_at
    """, (order_id,))

    details = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('admin/supplier_order_details.html',
                           token=token,
                           order=order,
                           details=details)

# Функція експорту в Excel замовлень постачальнику
@app.route('/<token>/admin/supplier-orders/<int:order_id>/export')
def export_supplier_order(token, order_id):
    """Export supplier order details to Excel file"""
    logging.info(f"Starting export for supplier order {order_id}")

    if validate_token(token):
        try:
            conn = get_db_connection()
            cur = conn.cursor()

            # Get order info
            cur.execute("""
                SELECT 
                    so.id,
                    so.created_at,
                    so.status,
                    so.order_number,
                    s.name as supplier_name
                FROM supplier_orders so
                JOIN suppliers s ON so.supplier_id = s.id
                WHERE so.id = %s
            """, [order_id])
            order = cur.fetchone()

            if order:
                # Get order details
                cur.execute("""
                    SELECT article, quantity, tracking_code, created_at
                    FROM supplier_order_details
                    WHERE supplier_order_id = %s
                """, [order_id])
                items = cur.fetchall()

                # Create Excel workbook
                workbook = Workbook()
                sheet = workbook.active

                # Add headers
                sheet['A1'] = f'Supplier Order #{order["order_number"]}'
                sheet['A2'] = f'Supplier: {order["supplier_name"]}'
                sheet['A3'] = f'Status: {order["status"]}'
                sheet['A4'] = f'Created: {order["created_at"]}'

                # Add column headers
                sheet['A6'] = 'Article'
                sheet['B6'] = 'Quantity'
                sheet['C6'] = 'Tracking Code' 
                sheet['D6'] = 'Created At'

                # Add order items
                row = 7
                for item in items:
                    sheet[f'A{row}'] = item['article']
                    sheet[f'B{row}'] = item['quantity']
                    sheet[f'C{row}'] = item['tracking_code']
                    sheet[f'D{row}'] = item['created_at']
                    row += 1

                # Save to BytesIO buffer
                output = BytesIO()
                workbook.save(output)
                output.seek(0)

                # Return Excel file
                filename = f'supplier_order_{order["order_number"]}.xlsx'
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )

        except Exception as e:
            logging.error(f"Export failed: {str(e)}")
            flash("Export failed. Please try again.", "error")

        finally:
            cur.close()
            conn.close()

    return redirect(url_for('list_supplier_orders', token=token))


@app.route('/<token>/upload_file', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def upload_file(token):
    try:
        logging.debug(f"Upload File Called with token: {token}")
        # Очищення сесійних даних
        session['missing_articles'] = []
        session['items_without_table'] = []
        merged_articles = {}  # Словник для відстеження об'єднаних артикулів

        user_id = session.get('user_id')
        if not user_id:
            logging.error("User not authenticated.")
            flash("User not authenticated.", "error")
            return redirect(f'/{token}/')

        logging.info(f"User ID: {user_id} started file upload.")

        # Перевірка наявності файлу
        if 'file' not in request.files:
            logging.error("No file uploaded.")
            flash("No file uploaded.", "error")
            return redirect(f'/{token}/')

        file = request.files['file']
        logging.info(f"Uploaded file: {file.filename}")

        # Перевірка формату файлу
        if not file.filename.endswith('.xlsx'):
            logging.error("Invalid file format. Only .xlsx files are allowed.")
            flash("Invalid file format. Please upload an Excel file.", "error")
            return redirect(f'/{token}/')

        # Зчитування файлу
        try:
            df = pd.read_excel(file, header=None)
            logging.info(f"File read successfully. Shape: {df.shape}")
        except Exception as e:
            logging.error(f"Error reading Excel file: {e}", exc_info=True)
            flash("Error reading the file. Please check the format.", "error")
            return redirect(f'/{token}/')

        if df.shape[1] < 2:
            logging.error("Invalid file structure. Less than two columns found.")
            flash("Invalid file structure. Ensure the file has at least two columns.", "error")
            return redirect(f'/{token}/')

        # Заповнення відсутніх колонок
        for col in range(2, 4):
            if col >= df.shape[1]:
                df[col] = None

        logging.debug(f"Dataframe after preprocessing: {df.head()}")

        items_with_table = []
        items_without_table = []
        missing_articles = set()

        # Обробка даних з файлу
        with get_db_connection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT pl.table_name, pl.brand_id, b.name as brand_name 
                FROM price_lists pl 
                JOIN brands b ON pl.brand_id = b.id
            """)
            all_tables = [(row[0], row[1], row[2]) for row in cursor.fetchall()]
            logging.info(f"Fetched tables from price_lists with brands: {all_tables}")

            for index, row in df.iterrows():
                try:
                    article = str(row[0]).strip() if pd.notna(row[0]) else None
                    quantity = int(row[1]) if pd.notna(row[1]) and str(row[1]).isdigit() else None
                    table_name = str(row[2]).strip() if pd.notna(row[2]) else None
                    comment = str(row[3]).strip() if pd.notna(row[3]) else None

                    if not article or not quantity:
                        logging.warning(f"Skipped invalid row at index {index}: {row.tolist()}")
                        continue

                    logging.debug(
                        f"Processing article: {article}, quantity: {quantity}, table: {table_name}, comment: {comment}")

                    if table_name:
                        table_info = next((t for t in all_tables if t[0] == table_name), None)
                        if not table_info:
                            logging.warning(f"Invalid table name '{table_name}' for article {article}.")
                            missing_articles.add(article)
                            continue
                        table_name, brand_id, brand_name = table_info

                        cursor.execute(f"SELECT price FROM {table_name} WHERE article = %s", (article,))
                        result = cursor.fetchone()

                        if result:
                            price = result[0]
                            items_with_table.append((article, price, table_name, quantity, comment, brand_id))
                            logging.info(f"Article {article} found in {table_name} with price {price}.")
                        else:
                            missing_articles.add(article)
                            logging.warning(f"Article {article} not found in {table_name}. Skipping.")
                    else:
                        matching_tables = []
                        for table, brand_id, brand_name in all_tables:
                            cursor.execute(f"SELECT price FROM {table} WHERE article = %s", (article,))
                            if cursor.fetchone():
                                matching_tables.append((table, brand_id, brand_name))

                        if matching_tables:
                            if len(matching_tables) == 1:
                                table, brand_id, brand_name = matching_tables[0]
                                cursor.execute(f"SELECT price FROM {table} WHERE article = %s", (article,))
                                price = cursor.fetchone()[0]
                                items_with_table.append((article, price, table, quantity, comment, brand_id))
                                logging.info(
                                    f"Article {article} automatically added from single table {table}")
                            else:
                                logging.info(f"Article {article} found in multiple tables: {matching_tables}")
                                items_without_table.append((article, quantity, matching_tables, comment))
                        else:
                            missing_articles.add(article)
                            logging.warning(f"Article {article} not found in any table.")
                except Exception as e:
                    logging.error(f"Error processing row at index {index}: {row.tolist()} - {e}", exc_info=True)
                    continue

            logging.debug(f"Items with table: {items_with_table}")
            logging.debug(f"Items without table: {items_without_table}")
            logging.debug(f"Missing articles: {missing_articles}")

            # Додавання товарів з визначеними таблицями до кошика
            if items_with_table:
                for article, base_price, table_name, quantity, comment, brand_id in items_with_table:
                    # Перевіряємо наявність в корзині перед додаванням
                    cursor.execute("""
                        SELECT quantity FROM cart 
                        WHERE user_id = %s AND article = %s AND table_name = %s
                    """, (user_id, article, table_name))
                    old_quantity = cursor.fetchone()

                    if old_quantity:
                        old_qty = old_quantity[0]
                        new_qty = old_qty + quantity
                        merged_articles[article] = [old_qty, new_qty]

                    role = session.get('role')
                    markup_percentage = get_markup_by_role(role)
                    final_price = calculate_price(base_price, markup_percentage)

                    cursor.execute("""
                        INSERT INTO cart (user_id, article, table_name, quantity, base_price, final_price, comment, brand_id, added_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON CONFLICT (user_id, article, table_name) 
                        DO UPDATE SET 
                            quantity = cart.quantity + EXCLUDED.quantity,
                            final_price = EXCLUDED.final_price,
                            comment = EXCLUDED.comment,
                            brand_id = EXCLUDED.brand_id
                    """, (user_id, article, table_name, quantity, base_price, final_price, comment, brand_id))
                    logging.info(f"Article {article} added/updated in cart")

            conn.commit()

            # Збереження даних в сесії та обробка результатів
            session['items_without_table'] = items_without_table
            session['missing_articles'] = list(missing_articles)

            # Додаємо повідомлення про об'єднані артикули
            if merged_articles:
                merge_msg = "Merged articles: " + ", ".join(
                    f"{art} ({old} + {new-old} = {new})"
                    for art, (old, new) in merged_articles.items()
                )
                flash(merge_msg, "info")

            if items_without_table:
                flash(f"{len(items_without_table)} articles need table selection.", "warning")
                return redirect(url_for('intermediate_results', token=token))

            if missing_articles:
                flash(f"The following articles were not found: {', '.join(missing_articles)}", "warning")

            flash(f"File processed successfully. {len(items_with_table)} items added to cart.", "success")
            return redirect(url_for('cart', token=token))

    except Exception as e:
        logging.error(f"Error in upload_file: {e}", exc_info=True)
        flash("An error occurred during file upload. Please try again.", "error")
        return redirect(f'/{token}/')



@app.route('/<token>/intermediate_results', methods=['GET', 'POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def intermediate_results(token):
    """
    Обробляє статті без таблиці, надає можливість вибрати таблиці із підтримкою brand_id,
    а потім додає вибрані статті до кошика.
    """
    logging.debug(f"Intermediate Results Called with token: {token}")
    try:
        # Отримання ідентифікатора користувача
        user_id = session.get('user_id')
        if not user_id:
            logging.error("Користувач не автентифікований.")
            flash(_("User not authenticated"), "error")
            return redirect(f'/{token}/')

        # Отримання націнки користувача
        user_markup = get_markup_percentage(user_id)
        logging.debug(f"Markup percentage for user_id={user_id}: {user_markup}%")

        if request.method == 'POST':
            logging.info("Обробка вибору таблиці для статей користувача.")
            # Отримання виборів користувача з форми
            # Очікується формат: "final_price|table_name|brand_id"
            user_selections = {}
            for key, value in request.form.items():
                if key.startswith('table_'):
                    article = key.split('_', 1)[1]
                    parts = value.split('|')
                    if len(parts) == 3:
                        try:
                            final_price = Decimal(parts[0])
                        except Exception:
                            final_price = Decimal('0')
                        table_name = parts[1]
                        brand_id = parts[2]
                        user_selections[article] = {
                            'final_price': final_price,
                            'table_name': table_name,
                            'brand_id': brand_id
                        }
            logging.debug(f"Вибір користувача: {user_selections}")

            items_without_table = session.get('items_without_table', [])
            missing_articles = session.get('missing_articles', [])
            added_to_cart = []

            with get_db_connection() as conn:
                cursor = conn.cursor()
                for item in items_without_table:
                    # Припускаємо, що item – кортеж з елементів: (article, quantity, available_tables, comment)
                    article = item[0]
                    if article in user_selections:
                        selection = user_selections[article]
                        final_price = selection['final_price']
                        table_name = selection['table_name']
                        brand_id = selection['brand_id']
                        quantity = item[1]
                        comment = item[3] if len(item) > 3 else None
                        # Вставка або оновлення статті в кошик з урахуванням brand_id
                        cursor.execute("""
                            INSERT INTO cart (user_id, article, table_name, quantity, base_price, final_price, comment, brand_id, added_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                            ON CONFLICT (user_id, article, table_name)
                            DO UPDATE SET 
                                quantity = cart.quantity + EXCLUDED.quantity,
                                final_price = EXCLUDED.final_price,
                                comment = EXCLUDED.comment,
                                brand_id = EXCLUDED.brand_id
                        """, (user_id, article, table_name, quantity, 0, final_price, comment, brand_id))
                        added_to_cart.append(article)
                        logging.info(f"Додано статтю {article} до кошика з таблицею {table_name} і brand_id {brand_id}.")
                conn.commit()

            # Оновлення сесії: видалення оброблених статей
            session['items_without_table'] = [item for item in items_without_table if item[0] not in added_to_cart]
            session['missing_articles'] = list(set(missing_articles))
            logging.debug(f"Оновлено сесію. Відсутні статті: {session['missing_articles']}")

            if session['items_without_table']:
                flash(_("Some articles still need table selection."), "warning")
            else:
                flash(_("All selected articles have been added to your cart."), "success")
            return redirect(url_for('cart', token=token))

        else:
            # GET: Відображення проміжних результатів
            items_without_table = session.get('items_without_table', [])
            missing_articles = session.get('missing_articles', [])
            logging.debug(f"Відображення проміжних результатів: items_without_table={len(items_without_table)}, missing_articles={len(missing_articles)}")

            enriched_items = []
            with get_db_connection() as conn:
                cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                # Формування доповнених даних для відображення вибору таблиці з підтримкою brand_id
                for article, quantity, valid_tables, comment in items_without_table:
                    item_prices = []
                    for table, brand_id, brand_name in valid_tables:
                        cursor.execute(
                            f"SELECT price FROM {table} WHERE article = %s",
                            (article,)
                        )
                        result = cursor.fetchone()
                        if result:
                            final_price = round(Decimal(result[0]) * (1 + user_markup / 100), 2)

                            item_prices.append({
                                'table': table,
                                'final_price': final_price,
                                'brand_id': brand_id,
                                'brand_name': brand_name  # Додаємо назву бренду
                            })
                    enriched_items.append({
                        'article': article,
                        'quantity': quantity,
                        'prices': item_prices,
                        'comment': comment
                    })
                cursor.close()
            return render_template(
                'user/search/intermediate.html',
                token=token,
                items_without_table=enriched_items,
                missing_articles=missing_articles
            )

    except Exception as e:
        logging.error(f"Помилка в intermediate_results: {e}", exc_info=True)
        flash(_("An error occurred while processing your selection. Please try again."), "error")
        return redirect(f'/{token}/')

# Кнопка підтвердження співпадінь аналізу інвойсу
@app.route('/<token>/admin/accept-matches/<int:invoice_id>', methods=['POST'])
@requires_token_and_roles('admin')
def accept_matches(token, invoice_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # 1. Оновлення повних співпадінь
            cursor.execute("""
                UPDATE order_details od
                SET status = 'invoice_received'
                FROM supplier_order_details sod
                JOIN invoice_details id ON sod.tracking_code = id.tracking_code
                WHERE sod.order_details_id = od.id
                AND id.invoice_id = %s
                AND od.status = 'ordered_supplier'
                AND id.quantity = od.quantity
            """, (invoice_id,))

            # 2. Обробка часткових співпадінь без створення нових замовлень
            cursor.execute("""
                WITH partial_matches AS (
                    SELECT 
                        od.id as order_detail_id,
                        od.order_id,
                        od.article,
                        od.quantity as order_quantity,
                        id.quantity as invoice_quantity,
                        od.base_price,
                        od.price,
                        od.comment,
                        od.table_name
                    FROM order_details od
                    JOIN supplier_order_details sod ON sod.order_details_id = od.id
                    JOIN invoice_details id ON sod.tracking_code = id.tracking_code
                    WHERE id.invoice_id = %s
                    AND od.status = 'ordered_supplier'
                    AND id.quantity < od.quantity
                )
                UPDATE order_details od
                SET 
                    quantity = pm.invoice_quantity,
                    status = 'invoice_received',
                    total_price = pm.price * pm.invoice_quantity
                FROM partial_matches pm
                WHERE od.id = pm.order_detail_id
            """, (invoice_id,))
            
            # 3. Оновлення статусу деталей інвойсу на 'processed'
            cursor.execute("""
                UPDATE invoice_details
                SET status = 'processed'
                WHERE invoice_id = %s
            """, (invoice_id,))
            
            # 4. Оновлення статусу інвойсу на 'accepted'
            cursor.execute("""
                UPDATE supplier_invoices
                SET status = 'accepted'
                WHERE id = %s
            """, (invoice_id,))
            
            conn.commit()
            flash("Matches accepted successfully", "success")
            
    except Exception as e:
        logging.error(f"Error accepting matches: {e}")
        flash("Error processing matches", "error")
        
    return redirect(url_for('list_invoices', token=token))


def get_markup_percentage(user_id):
    """
    Отримує відсоток націнки для користувача за його роллю.
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT r.markup_percentage
                FROM user_roles ur
                JOIN roles r ON ur.role_id = r.id
                WHERE ur.user_id = %s
            """, (user_id,))
            result = cursor.fetchone()
            return result[0] if result else 35  # За замовчуванням 35%
    except Exception as e:
        logging.error(f"Error fetching markup percentage for user_id={user_id}: {e}", exc_info=True)
        return 35  # Повертає стандартну націнку у разі помилки


@app.route('/<token>/admin/utilities')
@requires_token_and_roles('admin')
def utilities(token):
    return render_template('admin/utilities.html', token=token)


@app.route('/<token>/admin/news', methods=['GET'])
@requires_token_and_roles('admin')
def admin_news(token):
    """
    Відображення списку всіх новин в адмін-панелі
    """
    logging.info(f"Доступ до адмін-новин з токеном: {token}")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # Оновлений запит з вибіркою всіх мовних версій
        cursor.execute("""
            SELECT 
                id, 
                COALESCE(title_uk, title) as title_uk,
                COALESCE(title_en, title) as title_en,
                COALESCE(title_sk, title) as title_sk,
                created_at, 
                updated_at, 
                is_published
            FROM news
            ORDER BY created_at DESC
        """)
        news_list = cursor.fetchall()
        logging.info(f"Отримано {len(news_list)} новин")

        return render_template('admin/news/admin_news.html', news_list=news_list, token=token)

    except Exception as e:
        logging.error(f"Помилка отримання новин: {e}")
        flash("Помилка завантаження новин", "error")
        return redirect(url_for('admin_dashboard', token=token))

    finally:
        cursor.close()
        conn.close()


@app.route('/<token>/admin/news/create', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def create_news(token):
    if request.method == 'POST':
        try:
            title_uk = request.form['title_uk']
            content_uk = request.form['content_uk']
            html_content_uk = request.form['html_content_uk']

            title_en = request.form['title_en']
            content_en = request.form['content_en']
            html_content_en = request.form['html_content_en']

            title_sk = request.form['title_sk']
            content_sk = request.form['content_sk']
            html_content_sk = request.form['html_content_sk']

            is_published = bool(request.form.get('is_published'))

            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO news (
                    title_uk, content_uk, html_content_uk,
                    title_en, content_en, html_content_en,
                    title_sk, content_sk, html_content_sk,
                    is_published
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                title_uk, content_uk, html_content_uk,
                title_en, content_en, html_content_en,
                title_sk, content_sk, html_content_sk,
                is_published
            ))

            news_id = cursor.fetchone()[0]
            conn.commit()

            flash(_("News created successfully!"), "success")
            return redirect(url_for('admin_news', token=token))

        except Exception as e:
            logging.error(f"Error creating news: {e}")
            flash(_("Error creating news"), "error")
            return redirect(url_for('create_news', token=token))
        finally:
            if 'conn' in locals():
                cursor.close()
                conn.close()

    return render_template('admin/news/create_news.html', token=token)



# Прийняти всі артикулі в замовленні користувача
@app.route('/<token>/admin/orders/<int:order_id>/accept-all', methods=['POST'])
@requires_token_and_roles('admin')
def accept_all_items(token, order_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Перевіряємо чи є позиції зі статусом new
            cursor.execute("""
                SELECT COUNT(*) 
                FROM order_details 
                WHERE order_id = %s 
                AND status = 'new'
            """, (order_id,))
            
            unprocessed_count = cursor.fetchone()[0]
            
            if unprocessed_count == 0:
                flash("All items are already processed", "warning")
                return redirect(url_for('admin_order_details', token=token, order_id=order_id))

            # Оновлюємо тільки позиції зі статусом new
            cursor.execute("""
                UPDATE order_details 
                SET status = 'accepted',
                    processed_at = NOW()
                WHERE order_id = %s 
                AND status = 'new'
            """, (order_id,))
            
            # Оновлюємо статус замовлення
            cursor.execute("""
                UPDATE orders
                SET status = 'accepted'
                WHERE id = %s
                AND NOT EXISTS (
                    SELECT 1
                    FROM order_details
                    WHERE order_id = %s
                    AND status != 'accepted'
                )
            """, (order_id, order_id))
            
            conn.commit()
            flash("All items accepted successfully", "success")
            
    except Exception as e:
        logging.error(f"Error accepting all items: {e}")
        flash("Error processing items", "error")
        
    return redirect(url_for('admin_order_details', token=token, order_id=order_id))


# Маршрут для відображення форми вибору користувача для створення відвантаження
@app.route('/<token>/admin/shipments/create', methods=['GET'])
@requires_token_and_roles('admin')
def create_shipment_select_user(token):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо список всіх користувачів
            cursor.execute("""
                SELECT id, username FROM users
            """)
            users = cursor.fetchall()

            return render_template(
                'admin/shipments/select_user.html',
                token=token,
                users=users
            )

    except Exception as e:
        logging.error(f"Error fetching users for shipment creation: {e}", exc_info=True)
        flash("Error loading users for shipment creation", "error")
        return redirect(url_for('admin_dashboard', token=token))


# Маршрут для відображення форми створення відвантаження
@app.route('/<token>/admin/shipments/create/<int:user_id>', methods=['GET'])
@requires_token_and_roles('admin')
def create_shipment_form(token, user_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо інформацію про замовлення та деталі замовлення для вибраного користувача
            cursor.execute("""
                SELECT * FROM orders WHERE user_id = %s
            """, (user_id,))
            orders = cursor.fetchall()

            available_items = []
            for order in orders:
                cursor.execute("""
                    SELECT 
                        od.id,
                        od.article,
                        od.quantity,
                        od.price,
                        od.total_price,
                        od.order_id
                    FROM order_details od
                    WHERE od.order_id = %s AND od.shipment_id IS NULL AND od.status = 'invoice_received'
                """, (order['id'],))
                order_details = cursor.fetchall()
                available_items.extend(order_details)

            return render_template(
                'admin/shipments/create_shipment.html',
                token=token,
                user_id=user_id,
                available_items=available_items,
                orders=orders  # Передаємо список замовлень у шаблон
            )

    except Exception as e:
        logging.error(f"Error fetching data for shipment creation: {e}", exc_info=True)
        flash("Error loading data for shipment creation", "error")
        return redirect(url_for('admin_dashboard', token=token))


# Маршрут для відображення списку відвантажень для користувачів
@app.route('/<token>/shipments', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def user_shipments(token):
    try:
        user_id = session.get('user_id')
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            cursor.execute("""
                SELECT 
                    s.id,
                    s.shipment_date,
                    s.status,
                    s.tracking_number,
                    o.id as order_id,
                    COUNT(od.id) as items_count,
                    SUM(od.total_price) as total_price
                FROM shipments s
                JOIN orders o ON s.order_id = o.id
                LEFT JOIN order_details od ON od.shipment_id = s.id
                WHERE o.user_id = %s
                GROUP BY s.id, s.shipment_date, s.status, s.tracking_number, o.id
                ORDER BY s.shipment_date DESC
            """, (user_id,))
            
            shipments = cursor.fetchall()

            return render_template(
                'user/shipments/list_shipments.html',
                token=token,
                shipments=shipments
            )

    except Exception as e:
        logging.error(f"Error fetching user shipments: {e}", exc_info=True)
        flash("Error loading shipments", "error")
        return redirect(url_for('token_index', token=token))


# Маршрут для перегляду деталей відправлення користувачем
@app.route('/<token>/shipments/<int:shipment_id>', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def view_user_shipment(token, shipment_id):
    try:
        user_id = session.get('user_id')
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо інформацію про відправлення
            cursor.execute("""
                SELECT 
                    s.*,
                    o.id as order_id
                FROM shipments s
                JOIN orders o ON s.order_id = o.id
                WHERE s.id = %s AND o.user_id = %s
            """, (shipment_id, user_id))
            
            shipment = cursor.fetchone()
            if not shipment:
                flash("Shipment not found", "error")
                return redirect(url_for('user_shipments', token=token))

            # Отримуємо деталі відправлення
            cursor.execute("""
                SELECT 
                    od.article,
                    od.quantity,
                    od.price,
                    od.total_price,
                    od.comment
                FROM order_details od
                WHERE od.shipment_id = %s
            """, (shipment_id,))
            
            details = cursor.fetchall()

            return render_template(
                'user/shipments/view_shipment.html',
                token=token,
                shipment=shipment,
                details=details
            )

    except Exception as e:
        logging.error(f"Error fetching shipment details: {e}", exc_info=True)
        flash("Error loading shipment details", "error")
        return redirect(url_for('user_shipments', token=token))


# Маршрут для обробки створення відвантаження
@app.route('/<token>/admin/shipments/create', methods=['POST'])
@requires_token_and_roles('admin')
def create_shipment(token):
    try:
        user_id = request.form.get('user_id')
        order_id = request.form.get('order_id')
        tracking_number = request.form.get('tracking_number')
        selected_details = request.form.getlist('selected_details')

        if not selected_details:
            flash("Please select items for shipment", "error")
            return redirect(url_for('create_shipment_form', token=token, user_id=user_id))

        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Створюємо запис про відвантаження
            cursor.execute("""
                INSERT INTO shipments (order_id, shipment_date, status, tracking_number)
                VALUES (%s, CURRENT_DATE, 'created', %s)
                RETURNING id
            """, (order_id, tracking_number))
            shipment_id = cursor.fetchone()[0]

            # Оновлюємо shipment_id для вибраних деталей замовлення
            for detail_id in selected_details:
                cursor.execute("""
                    UPDATE order_details 
                    SET shipment_id = %s 
                    WHERE id = %s AND status = 'invoice_received'
                """, (shipment_id, detail_id))

            conn.commit()
            flash("Shipment created successfully", "success")

    except Exception as e:
        logging.error(f"Error creating shipment: {e}", exc_info=True)
        flash("Error creating shipment", "error")

    return redirect(url_for('list_shipments', token=token))

# Маршрут для перегляду деталей відвантаження
@app.route('/<token>/admin/shipments/<int:shipment_id>', methods=['GET'])
@requires_token_and_roles('admin')
def view_shipment(token, shipment_id):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо інформацію про відвантаження
            cursor.execute("""
                SELECT 
                    s.*,
                    o.id as order_id,
                    u.username as user_name,
                    u.email as user_email
                FROM shipments s
                JOIN orders o ON s.order_id = o.id
                JOIN users u ON o.user_id = u.id
                WHERE s.id = %s
            """, (shipment_id,))
            
            shipment = cursor.fetchone()
            if not shipment:
                flash("Shipment not found", "error")
                return redirect(url_for('list_shipments', token=token))

            # Отримуємо деталі відвантаження
            cursor.execute("""
                SELECT 
                    od.id,
                    od.article,
                    od.quantity,
                    od.price,
                    od.total_price,
                    od.comment
                FROM order_details od
                WHERE od.shipment_id = %s
            """, (shipment_id,))
            
            details = cursor.fetchall()

            return render_template(
                'admin/shipments/view_shipment.html',
                token=token,
                shipment=shipment,
                details=details
            )

    except Exception as e:
        logging.error(f"Error fetching shipment details: {e}", exc_info=True)
        flash("Error loading shipment details", "error")
        return redirect(url_for('list_shipments', token=token))

# Маршрут для відображення списку відвантажень
@app.route('/<token>/admin/shipments', methods=['GET'])
@requires_token_and_roles('admin')
def list_shipments(token):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Отримуємо список всіх відвантажень
            cursor.execute("""
                SELECT 
                    s.id,
                    s.shipment_date,
                    s.status,
                    s.tracking_number,
                    o.id as order_id,
                    u.username as user_name
                FROM shipments s
                JOIN orders o ON s.order_id = o.id
                JOIN users u ON o.user_id = u.id
                ORDER BY s.shipment_date DESC
            """)
            shipments = cursor.fetchall()

            return render_template(
                'admin/shipments/list_shipments.html',
                token=token,
                shipments=shipments
            )

    except Exception as e:
        logging.error(f"Error fetching shipments: {e}", exc_info=True)
        flash("Error loading shipments", "error")
        return redirect(url_for('admin_dashboard', token=token))




@app.route('/<token>/admin/export-orders', methods=['GET'])
@requires_token_and_roles('admin')
def export_orders(token):
    try:
        wb = Workbook()

        # Create sheets for each supplier
        mercedes_sheet = wb.active
        mercedes_sheet.title = "Mercedes"
        bmw_sheet = wb.create_sheet("BMW")
        vag_sheet = wb.create_sheet("VAG")

        # Set headers for each sheet
        headers = ['Article', 'Total Quantity', 'Price List']
        for sheet in [mercedes_sheet, bmw_sheet, vag_sheet]:
            sheet.append(headers)

        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            query = """
                SELECT 
                    od.article,
                    SUM(od.quantity) as total_quantity,
                    od.table_name,
                    CASE 
                        WHEN od.table_name LIKE '%mercedes%' THEN 'Mercedes'
                        WHEN od.table_name LIKE '%bmw%' THEN 'BMW'
                        WHEN od.table_name LIKE '%vag%' THEN 'VAG'
                        ELSE 'Unknown'
                    END as supplier
                FROM order_details od
                GROUP BY od.article, od.table_name
                ORDER BY od.article;
            """

            cursor.execute(query)
            results = cursor.fetchall()

            for row in results:
                data = [row['article'], row['total_quantity'], row['table_name']]
                if 'mercedes' in row['table_name'].lower():
                    mercedes_sheet.append(data)
                elif 'bmw' in row['table_name'].lower():
                    bmw_sheet.append(data)
                elif 'vag' in row['table_name'].lower():
                    vag_sheet.append(data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='supplier_orders.xlsx'
        )

    except Exception as e:
        logging.error(f"Error exporting orders: {e}")
        flash("Error exporting orders", "error")
        return redirect(url_for('admin_dashboard', token=token))


@app.route('/api/update-supplier-statuses', methods=['POST'])
@requires_token_and_roles('admin')
def update_supplier_statuses():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        df = pd.read_excel(file)

        conn = get_db_connection()
        cursor = conn.cursor()

        for _, row in df.iterrows():
            cursor.execute("""
                UPDATE order_details 
                SET status = %s 
                WHERE article = %s AND order_id = %s
            """, (row['status'], row['article'], row['order_id']))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        logging.error(f"Error updating statuses: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/process-supplier-invoice', methods=['POST'])
@requires_token_and_roles('admin')
def process_supplier_invoice():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        df = pd.read_excel(file)

        conn = get_db_connection()
        cursor = conn.cursor()

        # Створюємо новий інвойс
        cursor.execute("""
            INSERT INTO supplier_invoices (invoice_number, supplier_id, status)
            VALUES (%s, %s, 'received')
            RETURNING id
        """, (df['invoice_number'].iloc[0], df['supplier_id'].iloc[0]))

        invoice_id = cursor.fetchone()[0]

        # Додаємо зв'язки замовлень з інвойсом
        for _, row in df.iterrows():
            cursor.execute("""
                INSERT INTO order_invoice_links 
                (order_id, invoice_id, article, quantity, status)
                VALUES (%s, %s, %s, %s, 'in_transit')
            """, (row['order_id'], invoice_id, row['article'], row['quantity']))

            # Оновлюємо статус замовлення
            cursor.execute("""
                UPDATE order_details 
                SET status = 'in_transit' 
                WHERE article = %s AND order_id = %s
            """, (row['article'], row['order_id']))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        logging.error(f"Error processing invoice: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# endpoint accept-match для підтвердження в інвойсах постачальника
@app.route('/<token>/api/accept-match', methods=['POST'])
@requires_token_and_roles('admin')
def accept_match():
    try:
        data = request.json
        invoice_detail_id = data.get('invoice_detail_id')
        order_detail_id = data.get('order_detail_id')
        
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE order_details 
                SET status = 'invoice_received'
                WHERE id = %s
            """, (order_detail_id,))
            
            conn.commit()
            
        return jsonify({'success': True})
    except Exception as e:
        logging.error(f"Error accepting match: {e}")
        return jsonify({'error': str(e)}), 500


# експорт результату аналізу інвойса
@app.route('/<token>/admin/invoice/<int:invoice_id>/export', methods=['GET'])
@requires_token_and_roles('admin')
def export_invoice_analysis(token, invoice_id):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Analysis"
        
        # Headers
        headers = ['Article', 'Invoice Qty', 'Order Qty', 'Tracking Code', 'Status', 'Details']
        ws.append(headers)
        
        # Add matched items
        for item in matching_items:
            ws.append([
                item['invoice_item']['article'],
                item['invoice_item']['quantity'],
                item['order_item']['order_quantity'],
                item['invoice_item']['tracking_code'],
                item['match_type'],
                f"Received: {item.get('received_qty', '')} Remaining: {item.get('remaining_qty', '')}"
            ])
            
        # Format and return Excel file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'invoice_analysis_{invoice_id}.xlsx'
        )
        
    except Exception as e:
        logging.error(f"Error exporting invoice analysis: {e}")
        flash("Error exporting analysis", "error")
        return redirect(url_for('analyze_invoice', token=token, invoice_id=invoice_id))


@app.route('/<token>/news')
@requires_token_and_roles('user', 'user_25', 'user_29')
def user_news(token):
    """
    Відображення списку новин для користувачів з підтримкою багатомовності
    """
    logging.info(f"Accessing user news with token: {token}")
    user_id = session.get('user_id')
    current_lang = session.get('language', 'uk')

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        cursor.execute(f"""
            SELECT 
                n.id,
                n.title_{current_lang} as title,
                n.content_{current_lang} as content,
                n.html_content_{current_lang} as html_content,
                n.created_at,
                CASE WHEN nr.id IS NOT NULL THEN true ELSE false END as is_read
            FROM news n
            LEFT JOIN news_reads nr ON nr.news_id = n.id AND nr.user_id = %s
            WHERE n.is_published = true
            ORDER BY n.created_at DESC
        """, (user_id,))

        news_list = cursor.fetchall()
        logging.info(f"Fetched {len(news_list)} news items for user {user_id} in language {current_lang}")

        return render_template('user/news/user_news.html',
                               news_list=news_list,
                               token=token,
                               current_lang=current_lang)

    except Exception as e:
        logging.error(f"Error fetching news for user: {e}")
        flash(_("Error loading news"), "error")
        return redirect(url_for('index', token=token))

    finally:
        cursor.close()
        conn.close()


@app.route('/<token>/news/<int:news_id>')
@requires_token_and_roles('user', 'user_25', 'user_29')
def get_news_details(token, news_id):
    """
    Отримання детальної інформації про новину
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        cursor.execute("""
            SELECT title, html_content, created_at
            FROM news
            WHERE id = %s AND is_published = true
        """, (news_id,))

        news = cursor.fetchone()
        if news:
            return jsonify({
                'title': news['title'],
                'html_content': news['html_content'],
                'created_at': news['created_at'].strftime('%d.%m.%Y %H:%M')
            })
        return jsonify({'error': _('News not found')}), 404

    except Exception as e:
        logging.error(f"Error fetching news details: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

def get_unread_news_count(user_id):
    """
    Підрахунок непрочитаних новин для користувача
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT COUNT(*) FROM news n
            LEFT JOIN news_reads nr ON nr.news_id = n.id AND nr.user_id = %s
            WHERE n.is_published = true AND nr.id IS NULL
        """, (user_id,))
        return cursor.fetchone()[0]
    finally:
        cursor.close()
        conn.close()

@app.context_processor
def inject_unread_news_count():
    if 'user_id' in session:
        return {'unread_news_count': get_unread_news_count(session['user_id'])}
    return {'unread_news_count': 0}


@app.route('/<token>/news/<int:news_id>/view', methods=['GET'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def view_news(token, news_id):
    """
    Відображення окремої новини з підтримкою багатомовності
    """
    # Отримуємо поточну мову користувача, за замовчуванням українська
    current_lang = session.get('language', 'uk')

    # Підключаємося до бази даних
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # Виконуємо запит з підтримкою старих та нових полів через COALESCE
        cursor.execute("""
            SELECT *, 
                COALESCE(title_uk, title) as title_uk,
                COALESCE(title_en, title) as title_en,
                COALESCE(title_sk, title) as title_sk,
                COALESCE(html_content_uk, html_content) as html_content_uk,
                COALESCE(html_content_en, html_content) as html_content_en,
                COALESCE(html_content_sk, html_content) as html_content_sk,
                created_at
            FROM news 
            WHERE id = %s
        """, (news_id,))

        news = cursor.fetchone()

        if not news:
            flash("Новину не знайдено", "error")
            return redirect(url_for('user_news', token=token))

        # Передаємо дані в шаблон
        return render_template('view_news.html',
                               news=news,
                               token=token,
                               news_id=news_id,
                               current_lang=current_lang)

    except Exception as e:
        logging.error(f"Помилка при відображенні новини {news_id}: {e}")
        flash("Помилка при завантаженні новини", "error")
        return redirect(url_for('user_news', token=token))

    finally:
        cursor.close()
        conn.close()


@app.route('/<token>/news/<int:news_id>/mark-read', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def mark_news_read(token, news_id):
    """
    Окремий endpoint для позначення новини як прочитаної
    """
    user_id = session.get('user_id')
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO news_reads (user_id, news_id, read_at)
        VALUES (%s, %s, NOW())
        ON CONFLICT (user_id, news_id) DO NOTHING
    """, (user_id, news_id))
    
    conn.commit()
    return jsonify({'success': True})

@app.route('/<token>/news/<int:news_id>/set-read', methods=['POST'])
@app.route('/<token>/news/<int:news_id>/set-read', methods=['POST'])
@requires_token_and_roles('user', 'user_25', 'user_29')
def set_news_as_read(token, news_id):
    """
    Endpoint for marking news as read
    """
    user_id = session.get('user_id')
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO news_reads (user_id, news_id, read_at)
        VALUES (%s, %s, NOW())
        ON CONFLICT (user_id, news_id) DO NOTHING
    """, (user_id, news_id))

    conn.commit()
    return jsonify({'success': True})


@app.route('/<token>/admin/news/<int:news_id>/edit', methods=['GET', 'POST'])
@requires_token_and_roles('admin')
def edit_news(token, news_id):
    """
    Редагування багатомовної новини
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        if request.method == 'POST':
            # Отримуємо дані для кожної мови
            data = {
                'uk': {
                    'title': request.form['title_uk'],
                    'content': request.form['content_uk'],
                    'html_content': request.form['html_content_uk']
                },
                'en': {
                    'title': request.form['title_en'],
                    'content': request.form['content_en'],
                    'html_content': request.form['html_content_en']
                },
                'sk': {
                    'title': request.form['title_sk'],
                    'content': request.form['content_sk'],
                    'html_content': request.form['html_content_sk']
                }
            }

            is_published = 'is_published' in request.form

            cursor.execute("""
                UPDATE news 
                SET title_uk = %s, content_uk = %s, html_content_uk = %s,
                    title_en = %s, content_en = %s, html_content_en = %s,
                    title_sk = %s, content_sk = %s, html_content_sk = %s,
                    is_published = %s, updated_at = NOW()
                WHERE id = %s
            """, (
                data['uk']['title'], data['uk']['content'], data['uk']['html_content'],
                data['en']['title'], data['en']['content'], data['en']['html_content'],
                data['sk']['title'], data['sk']['content'], data['sk']['html_content'],
                is_published, news_id
            ))

            conn.commit()
            flash("Новину успішно оновлено!", "success")
            return redirect(url_for('admin_news', token=token))

        cursor.execute("SELECT * FROM news WHERE id = %s", (news_id,))
        news = cursor.fetchone()

        if not news:
            flash("Новину не знайдено", "error")
            return redirect(url_for('admin_news', token=token))

        return render_template('admin/news/edit_news.html', news=news, token=token)

    except Exception as e:
        logging.error(f"Помилка редагування новини: {e}")
        flash("Помилка оновлення новини", "error")
        return redirect(url_for('admin_news', token=token))

    finally:
        cursor.close()
        conn.close()


@app.route('/ping', methods=['GET'])
def ping():
    logging.info("ping function called.")  
    return "OK", 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"Starting server on port {port}...")
    app.run(host='0.0.0.0', port=port)

